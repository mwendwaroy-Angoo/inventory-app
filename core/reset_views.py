"""
Reset Sales & Analytics — owner-only. Permanently wipes a business's sales/
transaction/analytics history (Transactions, Receipts, Tabs, Shifts, keg/
produce/kitchen envelopes, debt, expenses, etc.) while keeping the business
itself, staff, and item catalog (Items/Presets/Stores) fully intact. Built
for businesses that need a genuine clean slate — see CLAUDE.md sprint log
for the full design rationale (why this is a hard delete, not a soft
cutover filter, and why stock balances are zeroed rather than frozen from
the pre-reset computed values).
"""
import logging

from django.contrib import messages
from django.db import transaction as db_transaction
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.utils import timezone
from django.utils.translation import gettext as _
from django.views.decorators.http import require_POST
import openpyxl

from .models import (
    Item, Transaction, Receipt, BarTab, Shift, KegBarrel, ProduceBunch,
    KitchenBatch, KitchenConsumableLog, Payment, CustomerDebtPayment,
    Customer, PerformerSession, StockRequest, BusinessExpense, PettyCash,
    SalaryPayment, SalaryDeduction, StockTake, Order, Forecast, Notification,
    TableOrder, BarCupLog, ProduceOverhead, ItemSaleApproval,
    PendingTransactionPrompt, CapitalInvestment, SalesResetLog,
)
from .views import get_user_profile, owner_required, owner_or_manager_required

logger = logging.getLogger(__name__)

# Models wiped by a reset — each has its own direct `business` FK and is
# deleted with one bulk `.filter(business=business).delete()` call (no
# Python row-loop). Some cascade child rows automatically and don't need
# their own entry: BarTabEntry (via BarTab), ShiftStockCount (via Shift),
# KegWeightReading (via KegBarrel), StockVarianceQuery (via StockTake),
# OrderLine (via Order), TableOrderItem (via TableOrder), WriteOffRequest
# (OneToOne to Transaction). KitchenConsumableLog has its OWN direct
# business FK and does NOT cascade from KitchenBatch, so it needs its own
# call. Notification has no business FK at all — scoped via the owning
# user's profile instead.
_WIPE_MODELS = [
    ('Transaction', Transaction),
    ('Receipt', Receipt),
    ('BarTab', BarTab),
    ('Shift', Shift),
    ('KegBarrel', KegBarrel),
    ('ProduceBunch', ProduceBunch),
    ('KitchenBatch', KitchenBatch),
    ('KitchenConsumableLog', KitchenConsumableLog),
    ('Payment', Payment),
    ('CustomerDebtPayment', CustomerDebtPayment),
    ('Customer', Customer),
    ('PerformerSession', PerformerSession),
    ('StockRequest', StockRequest),
    ('BusinessExpense', BusinessExpense),
    ('PettyCash', PettyCash),
    ('SalaryPayment', SalaryPayment),
    ('SalaryDeduction', SalaryDeduction),
    ('StockTake', StockTake),
    ('Order', Order),
    ('Forecast', Forecast),
    ('TableOrder', TableOrder),
    ('BarCupLog', BarCupLog),
    ('ProduceOverhead', ProduceOverhead),
    ('ItemSaleApproval', ItemSaleApproval),
    ('PendingTransactionPrompt', PendingTransactionPrompt),
]

# Explicitly NOT wiped — marketplace/cross-business models where "this
# business's" rows are tangled with another business's (Feedback has both
# from_business/to_business; supplier-chain models link two different
# businesses). Wiping one side's rows for a single-business reset would
# leave the other business's copy of the interaction meaningless/orphaned.
# Deliberately out of scope — listed here only so the omission reads as a
# decision, not an oversight, for the next person auditing this file.
# (Feedback, SupplierRelationship, SupplierBid, SupplierApplication,
#  ProcurementRequest, PurchaseOrder/PurchaseOrderLine/GoodsReceipt/
#  GoodsReceiptLine)

# Kept — structural/config, not sales/analytics history: Item,
# ItemPortionPreset, Store, Category, the Business row + settings,
# UserProfile, RecurringExpense (rule definitions — instances land in
# BusinessExpense, which IS wiped), Performer (roster — PerformerSession,
# the actual session history, IS wiped), RevenueTarget (a goal, not a
# record). CapitalInvestment is also kept by default (a durable business
# fact, like pre_app_cumulative_profit) — its total is shown on the intro
# page so the owner can see it wasn't touched.


def _counts_for_business(business):
    counts = {label: model.objects.filter(business=business).count() for label, model in _WIPE_MODELS}
    counts['Notification'] = Notification.objects.filter(user__userprofile__business=business).count()
    return counts


def _backup_session_key(business_id):
    return f'reset_backup_ready_{business_id}'


@owner_required
def reset_sales_intro(request):
    up = get_user_profile(request)
    business = up.business
    counts = _counts_for_business(business)
    capital_total = CapitalInvestment.objects.filter(business=business).count()
    backup_ready = bool(request.session.get(_backup_session_key(business.id)))
    return render(request, 'core/reset_sales_intro.html', {
        'business': business,
        'counts': counts,
        'total_rows': sum(counts.values()),
        'capital_investment_count': capital_total,
        'backup_ready': backup_ready,
    })


@owner_required
def reset_sales_backup_download(request):
    """Downloads a full workbook of everything about to be deleted, one sheet
    per model, before the reset can be confirmed. Sets the session flag that
    unlocks Step 2 — the confirm form itself re-checks this server-side, so
    the flag is a UX sequencing aid only, not the real safety mechanism."""
    up = get_user_profile(request)
    business = up.business

    wb = openpyxl.Workbook()
    summary_ws = wb.active
    summary_ws.title = 'Summary'
    summary_ws.append(['Business', business.name])
    summary_ws.append(['Backup generated', timezone.now().strftime('%Y-%m-%d %H:%M')])
    summary_ws.append([])
    summary_ws.append(['Model', 'Row count'])
    counts = _counts_for_business(business)
    for label, count in counts.items():
        summary_ws.append([label, count])

    for label, model in _WIPE_MODELS:
        ws = wb.create_sheet(title=label[:31])  # Excel sheet-name length limit
        field_names = [f.name for f in model._meta.fields]
        ws.append(field_names)
        qs = model.objects.filter(business=business).order_by('pk')
        for obj in qs.iterator(chunk_size=500):
            ws.append([str(getattr(obj, f, '')) for f in field_names])

    notif_ws = wb.create_sheet(title='Notification')
    notif_ws.append(['id', 'user', 'title', 'message', 'notification_type', 'is_read', 'created_at'])
    notif_qs = Notification.objects.filter(
        user__userprofile__business=business
    ).select_related('user').order_by('pk')
    for n in notif_qs.iterator(chunk_size=500):
        notif_ws.append([n.id, n.user.username, n.title, n.message, n.notification_type, n.is_read, str(n.created_at)])

    filename = f'reset_backup_{business.id}_{timezone.now():%Y%m%d_%H%M%S}.xlsx'
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)

    request.session[_backup_session_key(business.id)] = timezone.now().isoformat()
    return response


@owner_required
@require_POST
def reset_sales_confirm(request):
    up = get_user_profile(request)
    business = up.business

    if not request.session.get(_backup_session_key(business.id)):
        messages.error(request, _('Please download the backup first.'))
        return redirect('reset_sales_intro')

    confirm_text = request.POST.get('confirm_text', '').strip()
    if confirm_text != business.name:
        messages.error(request, _('Business name did not match exactly. Please type it exactly as shown.'))
        return redirect('reset_sales_intro')

    reason = request.POST.get('reason', '').strip()

    with db_transaction.atomic():
        counts_snapshot = _counts_for_business(business)

        SalesResetLog.objects.create(
            business=business,
            business_name_cache=business.name,
            performed_by=request.user,
            performed_by_username_cache=request.user.username,
            reason=reason,
            counts_snapshot=counts_snapshot,
        )

        for _label, model in _WIPE_MODELS:
            model.objects.filter(business=business).delete()
        Notification.objects.filter(user__userprofile__business=business).delete()

        # Zero out — never freeze the pre-reset computed balance (it reflects
        # the non-compliant period, not physical reality). The owner does a
        # real recount afterward via the fresh-count checklist.
        Item.objects.filter(business=business).update(opening_bin_balance=0, opening_physical=0)

    # Clear the session flag now that the reset actually happened.
    request.session.pop(_backup_session_key(business.id), None)

    messages.success(request, _('Sales & analytics history has been reset.'))
    return redirect('reset_sales_complete')


@owner_required
def reset_sales_complete(request):
    up = get_user_profile(request)
    latest = SalesResetLog.objects.filter(business=up.business).order_by('-created_at').first()
    return render(request, 'core/reset_sales_complete.html', {'reset_log': latest})


@owner_or_manager_required
def fresh_stock_count_checklist(request):
    up = get_user_profile(request)
    business = up.business

    latest_reset = SalesResetLog.objects.filter(business=business).order_by('-created_at').first()
    if not latest_reset:
        messages.info(request, _('No reset has been performed for this business yet.'))
        return redirect('stock_list')

    counted_item_ids = Transaction.objects.filter(
        business=business, date__gte=latest_reset.created_at.date(),
    ).values_list('item_id', flat=True)

    from django.db.models import Q

    # Only items that existed AT the time of the reset ever had anything to
    # reconcile — Item.created_at is null for items that predate this field
    # (2026-07-22), treated as "old enough" rather than excluded. Without
    # this, an item added well after the reset (never zeroed, never part of
    # the wiped history) trivially matches "no transaction since reset"
    # purely because it's brand new, and gets swept onto the checklist —
    # confirmed live: Roy added 4 new items and they showed up here needing
    # a "fresh count" that was never applicable to them.
    pending_items = (
        Item.objects.filter(business=business, is_keg=False, is_produce=False)
        .filter(Q(created_at__isnull=True) | Q(created_at__lte=latest_reset.created_at))
        .exclude(id__in=counted_item_ids)
        .order_by('description')
    )

    return render(request, 'core/fresh_stock_count.html', {
        'reset_log': latest_reset,
        'pending_items': pending_items,
        'pending_count': pending_items.count(),
    })


@owner_or_manager_required
@require_POST
def mark_item_recounted(request, item_id):
    """Handles the one gap left by reusing adjust_stock_balance as-is: an
    item that's genuinely still at 0 after a real physical count produces
    no_change=True there and creates no transaction, so it would never
    leave the fresh-count checklist. This creates a zero-qty marker
    transaction instead, using the exact same [ADJ] invoice_no convention
    adjust_stock_balance already uses so it's excluded from the "missing
    cost price" home-dashboard alert."""
    up = get_user_profile(request)
    business = up.business
    item = Item.objects.filter(id=item_id, business=business).first()
    if not item:
        return redirect('fresh_stock_count_checklist')

    Transaction.objects.create(
        business=business,
        item=item,
        type='Receipt',
        qty=0,
        date=timezone.localdate(),
        recorded_by=request.user,
        recipient='Fresh count confirmed (still zero) — post-reset',
        invoice_no='[ADJ]',
        payment_method='',
    )
    messages.success(request, _('%(item)s confirmed at zero stock.') % {'item': item.description})
    return redirect('fresh_stock_count_checklist')
