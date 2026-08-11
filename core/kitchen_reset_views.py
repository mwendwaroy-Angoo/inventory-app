"""
Kitchen Item Reset — owner/manager-only. Wipes sales/receiving history for
ONE kitchen item since a chosen cutoff date, so a business can start fresh
after a ledger has drifted beyond what a normal correction can untangle.

2026-08-11 live report (Roy, Monsoon Inn): Kuku's per-preset received-vs-
sold anchor tally (see kitchen_views.py's hidden_presets diagnostic) showed
a genuine, confirmed drift — 32.5 units sold against only 23 received —
traced to a real, deliberate design choice: kitchen_stock_receipt_delete()
only ever removes the KitchenStockReceipt/KitchenStockReceiptLine
bookkeeping rows, never the real stock-adding Transaction underneath (so
deleting a genuine duplicate receipt never accidentally un-receives real
stock) — but that also means a genuinely-duplicated delivery's phantom
units stay in the item's overall balance forever, invisible to the preset
tracker. Separately, Chipo's raw-material cost correction (edit_raw_
material_cost) only reaches a batch still OPEN at correction time, leaving
every already-closed "bucket preparation" from before the fix permanently
wrong. Rather than chase precise algorithmic repair of either (fragile,
unverifiable without a live physical count), Roy asked to wipe both items'
history since their own most recent receiving event and re-enter from the
staff's paper sales book with proper backdating.

Mirrors core/reset_views.py's proven "Reset Sales & Analytics" pattern —
backup workbook first, typed confirmation required, everything inside one
atomic transaction, logged via the same SalesResetLog model (business-wide
in that feature, reused here with reason/counts_snapshot describing the
item-scoped run instead) — just scoped to one item's own kitchen-selling
records instead of the whole business.

Deliberately does NOT try to precisely identify/delete the specific
duplicate transaction responsible for a balance discrepancy, or the exact
raw-material Draw transaction that funded a since-deleted batch — there is
no reliable FK linking either back to "the one at fault," and guessing via
fuzzy qty/date matching risks deleting the WRONG row. Instead: after this
reset, the owner does one real physical recount via the existing ⚖️
Rekebisha tool on the item (and on the raw-material source item, if any)
to true up the balance to what's actually on the shelf — the same
established mechanism this app already uses for every other "balance is
wrong for reasons that can't be reconstructed" scenario.
"""
import logging
from datetime import datetime, time as _time

from django.contrib import messages
from django.db import transaction as db_transaction
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_POST
import openpyxl

from .models import (
    Item, KitchenBatch, KitchenStockReceipt, KitchenStockReceiptLine,
    SalesResetLog, Transaction,
)
from .views import get_user_profile, owner_or_manager_required

logger = logging.getLogger(__name__)


def _backup_session_key(item_id):
    return f'kitchen_item_reset_backup_ready_{item_id}'


def _default_cutoff(item):
    """Best-guess starting cutoff: the item's own most recent receiving
    event (a KitchenStockReceipt line, for a plain portion item, or a
    KitchenBatch's own received_on, for a batch item). Falls back to
    today if the item has no receiving history through either mechanism
    yet — the owner can always type a different date."""
    if item.is_kitchen_batch:
        latest = (
            KitchenBatch.objects.filter(item=item)
            .order_by('-received_on').values_list('received_on', flat=True).first()
        )
    else:
        latest = (
            KitchenStockReceiptLine.objects.filter(item=item)
            .order_by('-receipt__received_on')
            .values_list('receipt__received_on', flat=True).first()
        )
    return latest or timezone.localdate()


def _scope_for_item(business, item, cutoff_date):
    """Everything this reset would touch for `item` since `cutoff_date`
    (a date, inclusive — the cutoff day itself counts). Returns a dict of
    querysets/counts/totals; never mutates anything."""
    cutoff_dt = timezone.make_aware(
        datetime.combine(cutoff_date, _time.min), timezone.get_current_timezone(),
    )

    if item.is_kitchen_batch:
        batches_qs = KitchenBatch.objects.filter(
            business=business, item=item, received_on__gte=cutoff_date,
        )
        batch_ids = list(batches_qs.values_list('id', flat=True))
        txns_qs = Transaction.objects.filter(
            business=business, kitchen_batch_id__in=batch_ids,
        )
        receipts_qs = KitchenStockReceipt.objects.none()
    else:
        batches_qs = KitchenBatch.objects.none()
        txns_qs = Transaction.objects.filter(
            business=business, item=item, created_at__gte=cutoff_dt,
        )
        receipt_ids = list(
            KitchenStockReceiptLine.objects.filter(
                item=item, receipt__business=business,
                receipt__received_on__gte=cutoff_date,
            ).values_list('receipt_id', flat=True).distinct()
        )
        receipts_qs = KitchenStockReceipt.objects.filter(id__in=receipt_ids)

    # A sale linked to a live BarTabEntry (Food Tab) has wider blast radius
    # than this item alone — deleting it would cascade-delete the tab entry
    # too (BarTabEntry.transaction is on_delete=CASCADE). Never silently
    # touched; surfaced separately so the owner can see if any exist.
    tab_linked_qs = txns_qs.filter(tab_entry__isnull=False)
    deletable_txns_qs = txns_qs.filter(tab_entry__isnull=True)

    total_revenue = sum((t.revenue() for t in deletable_txns_qs), 0.0)

    return {
        'cutoff_dt': cutoff_dt,
        'batches_qs': batches_qs,
        'receipts_qs': receipts_qs,
        'deletable_txns_qs': deletable_txns_qs,
        'tab_linked_qs': tab_linked_qs,
        'batch_count': batches_qs.count(),
        'receipt_count': receipts_qs.count(),
        'txn_count': deletable_txns_qs.count(),
        'tab_linked_count': tab_linked_qs.count(),
        'total_revenue': round(total_revenue, 2),
    }


@owner_or_manager_required
def kitchen_item_reset_intro(request, item_id):
    up = get_user_profile(request)
    business = up.business
    item = get_object_or_404(Item, id=item_id, store__business=business)

    cutoff_raw = request.GET.get('cutoff', '').strip()
    try:
        cutoff_date = datetime.strptime(cutoff_raw, '%Y-%m-%d').date() if cutoff_raw else _default_cutoff(item)
    except ValueError:
        cutoff_date = _default_cutoff(item)

    scope = _scope_for_item(business, item, cutoff_date)
    backup_ready = bool(request.session.get(_backup_session_key(item.id)))

    return render(request, 'core/kitchen_item_reset_intro.html', {
        'item': item,
        'cutoff_date': cutoff_date,
        'scope': scope,
        'current_balance': item.current_balance(),
        'backup_ready': backup_ready,
    })


def kitchen_item_reset_backup(request, item_id):
    up = get_user_profile(request)
    if not up or not getattr(up, 'is_owner_or_manager', False):
        messages.error(request, "Only business owners and managers can access this page.")
        return redirect('home')
    business = up.business
    item = get_object_or_404(Item, id=item_id, store__business=business)

    cutoff_raw = request.GET.get('cutoff', '').strip()
    try:
        cutoff_date = datetime.strptime(cutoff_raw, '%Y-%m-%d').date() if cutoff_raw else _default_cutoff(item)
    except ValueError:
        cutoff_date = _default_cutoff(item)
    scope = _scope_for_item(business, item, cutoff_date)

    wb = openpyxl.Workbook()
    summary_ws = wb.active
    summary_ws.title = 'Summary'
    summary_ws.append(['Item', item.description])
    summary_ws.append(['Cutoff (inclusive)', str(cutoff_date)])
    summary_ws.append(['Backup generated', timezone.now().strftime('%Y-%m-%d %H:%M')])
    summary_ws.append(['Current item balance', float(item.current_balance())])
    summary_ws.append(['Transactions to remove', scope['txn_count']])
    summary_ws.append(['Revenue being removed (KES)', scope['total_revenue']])
    summary_ws.append(['Tab-linked sales excluded (not touched)', scope['tab_linked_count']])

    txn_ws = wb.create_sheet(title='Transactions')
    txn_fields = ['id', 'type', 'qty', 'sale_amount', 'payment_method', 'recipient', 'invoice_no', 'created_at']
    txn_ws.append(txn_fields)
    for t in scope['deletable_txns_qs'].order_by('created_at'):
        txn_ws.append([str(getattr(t, f, '')) for f in txn_fields])

    if scope['receipt_count']:
        rcpt_ws = wb.create_sheet(title='Stock Receipts')
        rcpt_ws.append(['id', 'supplier', 'invoice_no', 'received_on', 'status'])
        for r in scope['receipts_qs']:
            rcpt_ws.append([r.id, r.supplier, r.invoice_no, str(r.received_on), r.status])

    if scope['batch_count']:
        batch_ws = wb.create_sheet(title='Kitchen Batches')
        batch_ws.append(['id', 'cost_total', 'revenue_collected', 'status', 'received_on'])
        for b in scope['batches_qs']:
            batch_ws.append([b.id, float(b.cost_total), float(b.revenue_collected), b.status, str(b.received_on)])

    filename = f'kitchen_item_reset_backup_{item.id}_{timezone.now():%Y%m%d_%H%M%S}.xlsx'
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)

    request.session[_backup_session_key(item.id)] = timezone.now().isoformat()
    return response


@owner_or_manager_required
@require_POST
def kitchen_item_reset_confirm(request, item_id):
    up = get_user_profile(request)
    business = up.business
    item = get_object_or_404(Item, id=item_id, store__business=business)

    cutoff_raw = request.POST.get('cutoff', '').strip()
    try:
        cutoff_date = datetime.strptime(cutoff_raw, '%Y-%m-%d').date()
    except ValueError:
        messages.error(request, 'Tarehe si sahihi.')
        return redirect('kitchen_item_reset_intro', item_id=item.id)

    if not request.session.get(_backup_session_key(item.id)):
        messages.error(request, 'Pakua nakala ya chelezo (backup) kwanza.')
        return redirect(f"{reverse('kitchen_item_reset_intro', args=[item.id])}?cutoff={cutoff_date}")

    confirm_text = request.POST.get('confirm_text', '').strip()
    if confirm_text != item.description:
        messages.error(request, 'Jina la bidhaa halikufanana kabisa. Andika tena kama lilivyoonyeshwa.')
        return redirect(f"{reverse('kitchen_item_reset_intro', args=[item.id])}?cutoff={cutoff_date}")

    reason = request.POST.get('reason', '').strip()

    scope = _scope_for_item(business, item, cutoff_date)
    with db_transaction.atomic():
        counts_snapshot = {
            'item': item.description,
            'cutoff': str(cutoff_date),
            'transactions_deleted': scope['txn_count'],
            'revenue_removed': scope['total_revenue'],
            'receipts_deleted': scope['receipt_count'],
            'batches_deleted': scope['batch_count'],
            'tab_linked_excluded': scope['tab_linked_count'],
        }
        SalesResetLog.objects.create(
            business=business,
            business_name_cache=business.name,
            performed_by=request.user,
            performed_by_username_cache=request.user.username,
            reason=f"Kitchen item reset: {item.description} — {reason}".strip(' —'),
            counts_snapshot=counts_snapshot,
        )
        scope['deletable_txns_qs'].delete()
        scope['receipts_qs'].delete()
        scope['batches_qs'].delete()

    request.session.pop(_backup_session_key(item.id), None)
    messages.success(request, f'{item.description} — historia ya mauzo tangu {cutoff_date} imefutwa.')
    return redirect('kitchen_item_reset_complete', item_id=item.id)


@owner_or_manager_required
def kitchen_item_reset_complete(request, item_id):
    up = get_user_profile(request)
    business = up.business
    item = get_object_or_404(Item, id=item_id, store__business=business)
    latest = (
        SalesResetLog.objects.filter(business=business, reason__icontains=f'Kitchen item reset: {item.description}')
        .order_by('-created_at').first()
    )
    return render(request, 'core/kitchen_item_reset_complete.html', {
        'item': item,
        'reset_log': latest,
        'current_balance': item.current_balance(),
    })
