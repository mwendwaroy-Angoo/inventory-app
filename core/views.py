import logging
from decimal import Decimal, InvalidOperation
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.db.models import Q, Sum
from django.views.decorators.http import require_POST
from django.utils.translation import gettext as _
from .models import (
    Item,
    Transaction,
    Store,
    Customer,
    PurchaseOrder,
    PurchaseOrderLine,
    GoodsReceipt,
    GoodsReceiptLine,
    ItemPortionPreset,
    Category,
    BarTab,
    BarTabEntry,
    Notification,
    StockRequest,
)
from .forms import (
    ItemForm,
    PurchaseOrderForm,
    PurchaseOrderLineForm,
    PurchaseOrderLineFormSet,
    GoodsReceiptForm,
    GoodsReceiptLineFormSet,
)
from core.forecast_engine import run_ets, run_regression
import openpyxl
from datetime import date, timedelta
import json
import os

# ── HELPERS ──────────────────────────────────────────────────────────────────


def get_user_profile(request):
    try:
        return request.user.userprofile
    except Exception:
        return None


def health_check(request):
    """Lightweight health-check endpoint for Render's preboot / load balancer."""
    from django.db import connection

    try:
        connection.ensure_connection()
        return HttpResponse("ok", content_type="text/plain", status=200)
    except Exception:
        return HttpResponse("error", content_type="text/plain", status=503)


def manifest_json(request):
    """Serve the Web App Manifest at /manifest.json with proper content type."""
    from django.conf import settings
    import json

    manifest_path = settings.BASE_DIR / "static" / "manifest.json"
    try:
        with open(manifest_path) as f:
            manifest = json.load(f)
        return JsonResponse(manifest, content_type="application/manifest+json")
    except (FileNotFoundError, json.JSONDecodeError):
        return JsonResponse({}, content_type="application/manifest+json")


def service_worker(request):
    """Serve the Service Worker at /sw.js with proper scope headers."""
    from django.conf import settings

    sw_path = settings.BASE_DIR / "static" / "sw.js"
    try:
        with open(sw_path, "r") as f:
            content = f.read()
        response = HttpResponse(content, content_type="application/javascript")
        response["Service-Worker-Allowed"] = "/"
        response["Cache-Control"] = "no-cache"
        return response
    except FileNotFoundError:
        return HttpResponse("", content_type="application/javascript")


def offline(request):
    return render(request, "offline.html")


def owner_required(view_func):
    from functools import wraps

    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        try:
            if not request.user.userprofile.is_owner:
                messages.error(request, _("Only business owners can access this page."))
                return redirect("stock_list")
        except Exception:
            return redirect("home")
        return view_func(request, *args, **kwargs)

    return wrapper


def owner_or_manager_required(view_func):
    from functools import wraps

    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        try:
            if not request.user.userprofile.is_owner_or_manager:
                messages.error(request, _("Only business owners and managers can access this page."))
                return redirect("stock_list")
        except Exception:
            return redirect("home")
        return view_func(request, *args, **kwargs)

    return wrapper


# ── HOME ─────────────────────────────────────────────────────────────────────


def _station_scope(up):
    """Return (show_bar, show_kitchen) for a UserProfile.

    Owner sees both. Kitchen staff default to kitchen only unless can_access_bar.
    Bar / general / waitress staff default to bar only unless can_access_kitchen.
    Used throughout home(), shift_history(), and any view that must respect the
    Station Scoping Principle (see CLAUDE.md).
    """
    if up.is_owner_or_manager:
        return True, True
    if up.is_kitchen_staff:
        return bool(getattr(up, 'can_access_bar', False)), True
    # bar / general / waitress
    return True, bool(getattr(up, 'can_access_kitchen', False))


def home(request):
    context = {"today": timezone.now().strftime("%B %d, %Y")}

    if request.user.is_authenticated:
        profile = getattr(request.user, "userprofile", None)
        if profile and profile.role == "rider":
            return redirect("rider_dashboard")
        if profile and profile.role == "supplier":
            return redirect("supplier_dashboard")

        try:
            user_profile = request.user.userprofile
            business = user_profile.business

            # Station scoping — determine what this staff member can see
            show_bar, show_kitchen = _station_scope(user_profile)
            context['show_bar']     = show_bar
            context['show_kitchen'] = show_kitchen

            # Scope item list to relevant station(s)
            _item_qs = Item.objects.filter(business=business)
            if show_bar and not show_kitchen:
                _item_qs = _item_qs.filter(store__is_kitchen=False)
            elif show_kitchen and not show_bar:
                _item_qs = _item_qs.filter(store__is_kitchen=True)
            all_items = _item_qs

            reorder_items = [item for item in all_items if item.needs_reorder()]
            low_stock_count = len(
                [
                    item
                    for item in all_items
                    if item.current_balance() <= item.reorder_level
                ]
            )
            reorder_count = len(reorder_items)

            _reorder_sorted = sorted(reorder_items, key=lambda x: x.current_balance())[:20]
            for _ri in _reorder_sorted:
                _bal = _ri.current_balance()
                _unit_raw = (_ri.unit or '').strip().upper()
                if _unit_raw == 'ML' and abs(float(_bal)) >= 1000:
                    _litres = float(_bal) / 1000
                    _ri.balance_display = f"{_litres:.2f}".rstrip('0').rstrip('.')
                    _ri.unit_display = 'L'
                else:
                    _bf = float(_bal)
                    if _bf == int(_bf):
                        _ri.balance_display = f"{int(_bf):,}"
                    else:
                        _ri.balance_display = f"{_bf:.2f}".rstrip('0').rstrip('.')
                    _ri.unit_display = _ri.unit or ''
            context.update(
                {
                    "total_items": all_items.count(),
                    "low_stock_count": low_stock_count,
                    "reorder_count": reorder_count,
                    "reorder_items": _reorder_sorted,
                }
            )

            # Items received in last 7 days with no cost price set (owner dashboard alert)
            try:
                from datetime import timedelta as _td
                seven_days_ago = timezone.now().date() - _td(days=7)
                recent_receipts_no_cost = Transaction.objects.filter(
                    business=business,
                    type='Receipt',
                    date__gte=seven_days_ago,
                    item__cost_price__isnull=True,
                ).exclude(invoice_no='[ADJ]').values('item__description', 'item__id').distinct()
                context['items_missing_cost_price'] = list(recent_receipts_no_cost)
                context['missing_cost_price_count'] = len(context['items_missing_cost_price'])
            except Exception:
                context['items_missing_cost_price'] = []
                context['missing_cost_price_count'] = 0

            # Petty cash pending review count
            try:
                from .models import PettyCash as _PettyCash
                context['pending_petty_cash_count'] = _PettyCash.objects.filter(
                    business=business, status='pending'
                ).count()
            except Exception:
                context['pending_petty_cash_count'] = 0

            # Expiry alerts
            try:
                from datetime import date as _date, timedelta as _td
                _today = _date.today()
                _soon  = _today + _td(days=7)
                _exp_qs = Transaction.objects.filter(
                    business=business,
                    type='Receipt',
                    expiry_date__isnull=False,
                ).values('item_id').distinct()
                _expired_ids  = Transaction.objects.filter(
                    business=business, type='Receipt',
                    expiry_date__lt=_today,
                ).values_list('item_id', flat=True).distinct()
                _expiring_ids = Transaction.objects.filter(
                    business=business, type='Receipt',
                    expiry_date__gte=_today, expiry_date__lte=_soon,
                ).values_list('item_id', flat=True).distinct()
                context['expired_count']  = len(set(_expired_ids))
                context['expiring_count'] = len(set(_expiring_ids))
            except Exception:
                context['expired_count']  = 0
                context['expiring_count'] = 0

            # Pending restock requests badge (owner + manager)
            if user_profile.is_owner_or_manager:
                try:
                    context['pending_restocks'] = StockRequest.objects.filter(
                        business=business,
                        status__in=[StockRequest.STATUS_PENDING, StockRequest.STATUS_ORDERED],
                    ).count()
                except Exception:
                    context['pending_restocks'] = 0
            else:
                context['pending_restocks'] = 0

            # Pending stock variance queries badge (owner + manager)
            if user_profile.is_owner_or_manager:
                try:
                    from core.models import StockVarianceQuery
                    context['pending_variances_count'] = StockVarianceQuery.objects.filter(
                        stock_take__business=business,
                        status__in=[StockVarianceQuery.PENDING, StockVarianceQuery.RESPONDED],
                    ).count()
                except Exception:
                    context['pending_variances_count'] = 0
            else:
                context['pending_variances_count'] = 0

            # Fresh Stock Count banner — only relevant right after a Reset
            # Sales & Analytics run.
            if user_profile.is_owner_or_manager:
                try:
                    from core.models import SalesResetLog
                    _latest_reset = SalesResetLog.objects.filter(
                        business=business
                    ).order_by('-created_at').first()
                    if _latest_reset:
                        _counted_ids = Transaction.objects.filter(
                            business=business, date__gte=_latest_reset.created_at.date(),
                        ).values_list('item_id', flat=True)
                        # Only items that existed at reset time ever had anything to
                        # reconcile — see reset_views.py:fresh_stock_count_checklist
                        # for the full rationale (duplicated count, same fix required
                        # here per this app's own "audit every surface" rule).
                        context['fresh_count_pending'] = Item.objects.filter(
                            business=business, is_keg=False, is_produce=False,
                        ).filter(
                            Q(created_at__isnull=True) | Q(created_at__lte=_latest_reset.created_at)
                        ).exclude(id__in=_counted_ids).count()
                    else:
                        context['fresh_count_pending'] = 0
                except Exception:
                    context['fresh_count_pending'] = 0
            else:
                context['fresh_count_pending'] = 0

            # Bar revenue — computed for any user who can see the bar station,
            # independent of the keg module (bar revenue = all non-kitchen Issue txns).
            try:
                if show_bar:
                    _bar_txns = Transaction.objects.filter(
                        business=business, type='Issue',
                        date=timezone.localdate(),
                        payment_method__in=['cash', 'mpesa'],
                        item__store__is_kitchen=False,
                    ).exclude(payment_method='void').select_related('item')
                    context['bar_today_revenue'] = sum(t.revenue() for t in _bar_txns)
            except Exception:
                context['bar_today_revenue'] = 0

            # Keg-specific context: tapped kegs, low-fill alerts, DJ/MC sessions
            try:
                from .business_profiles import get_profile as _get_profile
                if _get_profile(business).get('modules', {}).get('keg'):
                    from .models import KegBarrel as _KB
                    _tapped = list(_KB.objects.filter(business=business, status='TAPPED'))
                    _at_risk = [k for k in _tapped
                                if k.remaining_envelope() < float(k.target_revenue or 1) * 0.15]
                    context['kegs_tapped'] = len(_tapped)
                    context['kegs_at_risk_count'] = len(_at_risk)
                    # Active DJ/MC sessions for home dashboard timer widget
                    try:
                        from .models import PerformerSession as _PS
                        _dj_today = timezone.localdate()
                        _active_dj = list(
                            _PS.objects.filter(
                                business=business,
                                status__in=[_PS.STATUS_ACTIVE, _PS.STATUS_PENDING_CONFIRMATION],
                                date=_dj_today,
                            ).select_related('performer', 'second_performer')
                            .order_by('started_at', 'created_at')
                        )
                        for _sess in _active_dj:
                            if _sess.started_at:
                                _sess.started_at_epoch = int(_sess.started_at.timestamp())
                            else:
                                _sess.started_at_epoch = int(_sess.created_at.timestamp())
                        context['active_dj_sessions'] = _active_dj
                    except Exception:
                        context['active_dj_sessions'] = []
            except Exception:
                context['kegs_tapped'] = 0
                context['kegs_at_risk_count'] = 0
                context['active_dj_sessions'] = []

            # Kitchen-specific today revenue (separate from bar)
            _has_kitchen = getattr(business, 'has_kitchen', False)
            context['has_kitchen'] = _has_kitchen
            try:
                if _has_kitchen and show_kitchen:
                    _kitchen_txns = Transaction.objects.filter(
                        business=business, type='Issue',
                        date=timezone.localdate(),
                        payment_method__in=['cash', 'mpesa'],
                        item__store__is_kitchen=True,
                    ).exclude(payment_method='void').select_related('item')
                    context['kitchen_today_revenue'] = sum(t.revenue() for t in _kitchen_txns)
                else:
                    context['kitchen_today_revenue'] = 0
            except Exception:
                context['kitchen_today_revenue'] = 0

            # Active managers logged in today (shown to owner on dashboard)
            if user_profile.is_owner:
                try:
                    from accounts.models import UserProfile as _ManagerUP
                    _today_start = timezone.localtime(timezone.now()).replace(
                        hour=0, minute=0, second=0, microsecond=0
                    )
                    context['active_managers'] = list(
                        _ManagerUP.objects.filter(
                            business=business,
                            role='manager',
                            user__last_login__gte=_today_start,
                        ).select_related('user').order_by('user__last_login')
                    )
                except Exception:
                    context['active_managers'] = []
            else:
                context['active_managers'] = []

            # Recurring expense review nudge (owner + manager, once per period)
            if user_profile.is_owner_or_manager:
                try:
                    from .models import RecurringExpense as _RE
                    today_d = timezone.localdate()
                    _review_date = business.last_expense_review_date
                    _has_recurring = _RE.objects.filter(business=business, is_active=True).exists()
                    _due = False
                    if _has_recurring:
                        month_start = today_d.replace(day=1)
                        _due = (not _review_date) or (_review_date < month_start)
                    context['expense_review_due'] = _due
                except Exception:
                    context['expense_review_due'] = False
            else:
                context['expense_review_due'] = False

            # Revenue targets progress for dashboard widget
            try:
                from core.models import RevenueTarget, Store as _Store
                from datetime import date as _date
                _today = _date.today()
                _week_start = _today - timedelta(days=_today.weekday())
                _month_start = _today.replace(day=1)

                def _period_rev(start, end):
                    txns = Transaction.objects.filter(
                        business=business, type='Issue',
                        date__gte=start, date__lte=end,
                    ).exclude(payment_method='void').select_related('item')
                    # Scope revenue to the staff member's station
                    if show_kitchen and not show_bar:
                        txns = txns.filter(item__store__is_kitchen=True)
                    elif show_bar and not show_kitchen:
                        txns = txns.filter(item__store__is_kitchen=False)
                    return sum(t.revenue() for t in txns)

                def _get_target(ttype):
                    # Kitchen-only staff: prefer per-store kitchen target
                    if show_kitchen and not show_bar:
                        _ks = _Store.objects.filter(business=business, is_kitchen=True).first()
                        if _ks:
                            _kt = RevenueTarget.objects.filter(
                                business=business, target_type=ttype, store=_ks
                            ).first()
                            if _kt:
                                return float(_kt.amount)
                    # Owner, cross-access, or bar staff: use business-wide target
                    t = RevenueTarget.objects.filter(
                        business=business, target_type=ttype, store__isnull=True
                    ).first()
                    return float(t.amount) if t else 0

                def _build_target_data(actual, target):
                    actual = round(float(actual), 2)
                    target = float(target)
                    if target > 0:
                        pct = min(100, round((actual / target) * 100, 1))
                    else:
                        pct = 0
                    if pct >= 100:
                        color = '#6ee7b7'
                    elif pct >= 50:
                        color = '#fbbf24'
                    else:
                        color = '#f87171'
                    return {'actual': actual, 'target': target, 'pct': pct, 'color': color}

                context['revenue_targets'] = {
                    'daily':   _build_target_data(_period_rev(_today, _today),       _get_target('daily')),
                    'weekly':  _build_target_data(_period_rev(_week_start, _today),  _get_target('weekly')),
                    'monthly': _build_target_data(_period_rev(_month_start, _today), _get_target('monthly')),
                }
            except Exception:
                context['revenue_targets'] = None

        except Exception:
            context["error"] = _("Profile not found. Please contact support.")
    else:
        context["guest"] = True
        context["services"] = [
            (
                "📦",
                _("Inventory Management"),
                _(
                    "Track stock levels, costs, and reorder points in real time. Never run out of stock again."
                ),
            ),
            (
                "🛒",
                _("Online Marketplace"),
                _(
                    "Your own storefront where customers browse and order directly. No middleman."
                ),
            ),
            (
                "💳",
                _("M-Pesa Payments"),
                _(
                    "Accept payments via Lipa Na M-Pesa. Instant STK Push to your customers' phones."
                ),
            ),
            (
                "📱",
                _("USSD Access"),
                _(
                    "Record sales and check stock via USSD. Works on any phone, no internet needed."
                ),
            ),
            (
                "📊",
                _("Analytics Dashboard"),
                _(
                    "See your top products, revenue trends, and profit margins at a glance."
                ),
            ),
            (
                "👥",
                _("Staff Management"),
                _(
                    "Add staff, assign roles, and get notified when they log in or record transactions."
                ),
            ),
        ]
        context["faqs"] = [
            (
                _("Is Duka Mwecheche free?"),
                _(
                    "Yes! The platform is completely free for all businesses. You only pay standard M-Pesa transaction fees when accepting payments."
                ),
            ),
            (
                _("Do I need a smartphone?"),
                _(
                    "No. You can manage your stock via USSD on any basic phone. The web app works on smartphones and computers too."
                ),
            ),
            (
                _("How do customers find my shop?"),
                _(
                    "Once you register and add items with prices, your business appears on the Marketplace. Customers can search by location and product."
                ),
            ),
            (
                _("Is my data safe?"),
                _(
                    "Absolutely. Your data is stored securely on cloud servers with regular backups. Only you and your staff can access your business data."
                ),
            ),
            (
                _("Can I accept M-Pesa payments?"),
                _(
                    "Yes. We integrate with Safaricom's Daraja API. You'll need a Till or Paybill number from Safaricom to receive funds directly."
                ),
            ),
            (
                _("How do I add staff?"),
                _(
                    "Go to Manage -> Staff -> Add Staff. Staff members can record transactions but cannot access business settings or financial reports."
                ),
            ),
        ]

    return render(request, "core/home.html", context)


def dashboard_revenue_api(request):
    """Live today's revenue for the home dashboard hero tiles (bar + kitchen).

    No @login_required — unauthenticated polls return zeros so the JS poll
    never triggers a login redirect loop (same pattern as notifications_count).
    """
    if not request.user.is_authenticated:
        return JsonResponse({'bar_revenue': 0, 'kitchen_revenue': 0, 'has_kitchen': False})
    try:
        up = request.user.userprofile
        business = up.business
        show_bar, show_kitchen = _station_scope(up)
        today = timezone.localdate()
        bar_rev = 0
        kitchen_rev = 0
        if show_bar:
            _bar_txns = Transaction.objects.filter(
                business=business, type='Issue',
                date=today,
                payment_method__in=['cash', 'mpesa'],
                item__store__is_kitchen=False,
            ).exclude(payment_method='void').select_related('item')
            bar_rev = sum(t.revenue() for t in _bar_txns)
        if show_kitchen and getattr(business, 'has_kitchen', False):
            _kit_txns = Transaction.objects.filter(
                business=business, type='Issue',
                date=today,
                payment_method__in=['cash', 'mpesa'],
                item__store__is_kitchen=True,
            ).exclude(payment_method='void').select_related('item')
            kitchen_rev = sum(t.revenue() for t in _kit_txns)
        return JsonResponse({
            'bar_revenue': round(bar_rev, 0),
            'kitchen_revenue': round(kitchen_rev, 0),
            'has_kitchen': getattr(business, 'has_kitchen', False),
        })
    except Exception:
        return JsonResponse({'bar_revenue': 0, 'kitchen_revenue': 0, 'has_kitchen': False})


# ── STOCK LIST ────────────────────────────────────────────────────────────────


@login_required
def stock_list(request):
    user_profile = get_user_profile(request)
    if not user_profile:
        messages.error(request, _("No business profile found."))
        return redirect("home")

    _sl_show_bar, _sl_show_kitchen = _station_scope(user_profile)
    # ?station=kitchen param forces kitchen view (e.g. from home dashboard card links)
    _station_param = request.GET.get("station", "")
    if _station_param == "kitchen":
        _sl_show_bar, _sl_show_kitchen = False, True

    stores = Store.objects.filter(business=user_profile.business)
    if _sl_show_kitchen and not _sl_show_bar:
        stores = stores.filter(is_kitchen=True)
    elif _sl_show_bar and not _sl_show_kitchen:
        stores = stores.filter(is_kitchen=False)

    selected_store_id = request.GET.get("store")
    status_filter = request.GET.get("status")

    items = Item.objects.filter(store__business=user_profile.business).exclude(is_keg=True)
    if _sl_show_kitchen and not _sl_show_bar:
        items = items.filter(store__is_kitchen=True)
    elif _sl_show_bar and not _sl_show_kitchen:
        items = items.exclude(store__is_kitchen=True)
    else:
        # Owner / cross-access: show all (bar + kitchen, no keg)
        items = items
    items = items.order_by("material_no")

    if selected_store_id:
        try:
            selected_store_id = int(selected_store_id)
            items = items.filter(store_id=selected_store_id)
        except (ValueError, TypeError):
            pass

    all_items = list(items)

    if status_filter == "low_stock":
        all_items = [i for i in all_items if i.current_balance() <= i.reorder_level]
    elif status_filter == "reorder":
        all_items = [i for i in all_items if i.needs_reorder()]

    # Annotate each item with its earliest expiry date from Receipt batches
    from datetime import date as _date, timedelta as _td
    from django.db.models import Min as _Min
    today_d = _date.today()
    soon_d  = today_d + _td(days=7)

    expiry_qs = (
        Transaction.objects
        .filter(
            business=user_profile.business,
            type='Receipt',
            expiry_date__isnull=False,
            item__in=all_items,
        )
        .values('item_id')
        .annotate(earliest=_Min('expiry_date'))
    )
    expiry_map = {row['item_id']: row['earliest'] for row in expiry_qs}

    for item in all_items:
        exp = expiry_map.get(item.id)
        if exp is None:
            item.expiry_date   = None
            item.expiry_status = None
        elif exp < today_d:
            item.expiry_date   = exp
            item.expiry_status = 'EXPIRED'
        elif exp <= soon_d:
            item.expiry_date   = exp
            item.expiry_status = 'EXPIRING'
        else:
            item.expiry_date   = exp
            item.expiry_status = 'OK'

    # Annotate items with pending restock request flag
    _pending_restock_ids = set(
        StockRequest.objects.filter(
            business=user_profile.business,
            status__in=[StockRequest.STATUS_PENDING, StockRequest.STATUS_ORDERED],
            item__in=all_items,
        ).values_list('item_id', flat=True)
    )
    for item in all_items:
        item.has_pending_restock = item.id in _pending_restock_ids

    if status_filter == "expiring":
        all_items = [i for i in all_items if i.expiry_status in ('EXPIRED', 'EXPIRING')]

    # Fresh Stock Count banner — only relevant right after a Reset Sales &
    # Analytics run, only shown to owner/manager.
    fresh_count_pending = 0
    if user_profile.is_owner_or_manager:
        from .models import SalesResetLog
        _latest_reset = SalesResetLog.objects.filter(
            business=user_profile.business
        ).order_by('-created_at').first()
        if _latest_reset:
            _counted_ids = Transaction.objects.filter(
                business=user_profile.business, date__gte=_latest_reset.created_at.date(),
            ).values_list('item_id', flat=True)
            # Same fix as the home() banner and fresh_stock_count_checklist —
            # only items that existed at reset time ever had anything to
            # reconcile; a brand-new item trivially matches "no transaction
            # since reset" purely because it's new.
            fresh_count_pending = Item.objects.filter(
                business=user_profile.business, is_keg=False, is_produce=False,
            ).filter(
                Q(created_at__isnull=True) | Q(created_at__lte=_latest_reset.created_at)
            ).exclude(id__in=_counted_ids).count()

    context = {
        "items": all_items,
        "stores": stores,
        "selected_store": selected_store_id if selected_store_id else None,
        "status_filter": status_filter,
        "today": timezone.now().strftime("%B %d, %Y"),
        "is_owner": user_profile.is_owner_or_manager,
        "fresh_count_pending": fresh_count_pending,
    }
    return render(request, "core/stock_list.html", context)


@login_required
def expiring_items(request):
    from datetime import date as _date, timedelta as _td
    from django.db.models import Min as _Min

    user_profile = get_user_profile(request)
    if not user_profile:
        return redirect("home")

    business = user_profile.business
    today_d  = _date.today()
    soon_d   = today_d + _td(days=7)

    # One query: earliest expiry per item, for all items with any expiry set
    expiry_rows = (
        Transaction.objects
        .filter(business=business, type='Receipt', expiry_date__isnull=False)
        .values('item_id', 'item__description', 'item__unit', 'item__store__name')
        .annotate(earliest=_Min('expiry_date'))
        .order_by('earliest')
    )

    items_data = []
    for row in expiry_rows:
        exp = row['earliest']
        if exp < today_d:
            status = 'EXPIRED'
            days   = (today_d - exp).days
            days_label = f"Expired {days} day{'s' if days != 1 else ''} ago"
        elif exp <= soon_d:
            status = 'EXPIRING'
            days   = (exp - today_d).days
            days_label = f"Expires in {days} day{'s' if days != 1 else ''}" if days > 0 else "Expires today"
        else:
            status = 'OK'
            days   = (exp - today_d).days
            days_label = f"Expires in {days} days"

        try:
            from .models import Item as _Item
            item_obj = _Item.objects.get(id=row['item_id'])
            balance = item_obj.current_balance()
        except Exception:
            balance = '—'

        items_data.append({
            'item_id':   row['item_id'],
            'name':      row['item__description'],
            'unit':      row['item__unit'],
            'store':     row['item__store__name'],
            'expiry':    exp,
            'status':    status,
            'days_label': days_label,
            'balance':   balance,
        })

    return render(request, 'core/expiring_items.html', {
        'items_data': items_data,
        'today': today_d,
    })


# ── TRANSACTIONS ──────────────────────────────────────────────────────────────


@login_required
def add_transaction(request):
    user_profile = get_user_profile(request)
    if not user_profile:
        return redirect("home")

    stores = Store.objects.filter(business=user_profile.business)
    customers = Customer.objects.filter(business=user_profile.business)

    if request.method == "POST":
        is_quick = request.GET.get('quick') == '1'
        restock_resolved = False

        # Server-side double-submit backstop for the fast AJAX path only (Quick
        # Sell's "+📦 Pata Stok" modal — built precisely for restocking mid-shift
        # under time pressure, the same busy-counter conditions that motivated
        # every other idempotency fix in this app). The normal full-page form
        # already has page-reload friction against accidental resubmission, so
        # it's left untouched — see core/idempotency.py. (Quick-Sell-module
        # audit finding, 2026-07-19.)
        if is_quick:
            from core.idempotency import claim_checkout_token
            idem_token = (request.POST.get('idempotency_token') or '').strip()
            if not claim_checkout_token(user_profile.business_id, idem_token):
                return JsonResponse({'ok': False, 'error': 'Hii tayari imehifadhiwa.', 'duplicate': True}, status=409)

        # Shift gate: staff must have an open shift to write any transaction
        if not user_profile.is_owner_or_manager:
            from core.shift_views import get_active_staff_shift
            if get_active_staff_shift(user_profile, user_profile.business) is False:
                messages.error(
                    request,
                    'Fungua shift yako kwanza kabla ya kuingiza muamala.'
                )
                return redirect('add_transaction')

        item_id = request.POST["item"]
        trans_type = request.POST["type"]
        try:
            quantity = Decimal(request.POST.get('quantity', '0'))
        except InvalidOperation:
            quantity = Decimal('0')
        invoice_no = request.POST.get("invoice_no", "")
        recipient = request.POST.get("recipient", "")

        new_customer_name = request.POST.get("new_customer_name", "").strip()
        if new_customer_name and trans_type == "Issue":
            customer = Customer.objects.filter(
                business=user_profile.business, name=new_customer_name
            ).first()
            if customer is None:
                customer = Customer.objects.create(
                    business=user_profile.business,
                    name=new_customer_name,
                    phone=request.POST.get("new_customer_phone", ""),
                )
            recipient = customer.name

        # ── CREDIT DISCIPLINE GATE for add_transaction ───────────────────────
        if (trans_type == "Issue"
                and request.POST.get("payment_method", "cash") == "credit"
                and recipient):
            from core.credit_policy import evaluate_credit
            _cust_gate = Customer.objects.filter(
                business=user_profile.business, name=recipient
            ).first()
            if _cust_gate is None:
                _cust_gate = Customer.objects.create(
                    business=user_profile.business, name=recipient,
                    credit_approved=True,
                )
            _decision = evaluate_credit(user_profile.business, _cust_gate)
            if not _decision.allowed:
                messages.error(
                    request,
                    f'Deni haliwezi kutolewa: {_decision.reason} — '
                    'Badilisha njia ya malipo.'
                )
                return redirect('add_transaction')
        # ─────────────────────────────────────────────────────────────────────

        # Multi-tenancy: this was get_object_or_404(Item, id=item_id) with NO
        # business filter at all — any authenticated staff member of ANY
        # business could submit another business's item_id and write bogus
        # Receipt/Issue/Wastage transactions straight into a stranger's stock
        # records (corrupting their balances, P&L, and triggering false
        # restock/expiry alerts). Reachable via the normal Add Transaction
        # form AND Quick Sell's "+📦 Pata Stok" quick=1 AJAX path (Quick-
        # Sell-module audit finding, 2026-07-19 — the most severe gap found
        # in this audit series so far).
        item = get_object_or_404(Item, id=item_id, store__business=user_profile.business)

        # ── PRODUCE PRESET HANDLING ───────────────────────────────────────────
        preset_id = request.POST.get('preset_id', '').strip()
        if preset_id and item.is_produce:
            try:
                preset = ItemPortionPreset.objects.get(id=int(preset_id), item=item)
                quantity = preset.quantity_consumed
            except (ItemPortionPreset.DoesNotExist, ValueError):
                pass
        # ─────────────────────────────────────────────────────────────────────

        # ── RESTRICTED ITEM CHECK ─────────────────────────────────────────────
        can_override = getattr(user_profile, 'can_override_restrictions', False)
        if trans_type == 'Issue' and item.is_restricted and not user_profile.is_owner_or_manager and not can_override:
            reserved = item.restricted_quantity or 0
            balance_after = item.current_balance() - quantity
            needs_approval = reserved == 0 or balance_after < reserved
            if needs_approval:
                from core.restricted_items_views import _create_approval_request
                approval = _create_approval_request(
                    request, item, user_profile,
                    quantity=quantity,
                    recipient=recipient,
                    invoice_no=invoice_no,
                    payment_method=request.POST.get('payment_method', 'cash'),
                )
                return render(request, 'core/sale_approval_pending.html', {
                    'approval': approval,
                    'item': item,
                })
            # else: sale is within freely-sellable range — falls through below
        # ─────────────────────────────────────────────────────────────────────

        if trans_type in ("Issue", "Wastage"):
            # Both Issue and Wastage reduce stock — qty must be stored negative.
            # Only Issue is guarded against going below zero; Wastage is allowed
            # to exceed current stock (e.g. recording expired stock that was
            # never properly received), but is still negated so balance falls.
            if trans_type == "Issue" and item.current_balance() < quantity:
                messages.error(
                    request,
                    _(
                        "Not enough stock for %(item_description)s. Available: %(available)s %(unit)s, requested: %(requested)s."
                    )
                    % {
                        "item_description": item.description,
                        "available": item.current_balance(),
                        "unit": item.unit,
                        "requested": quantity,
                    },
                )
                return redirect("add_transaction")
            quantity = -quantity

        # Backdated timestamp (Option B offline sales)
        backdated_at = None
        backdated_raw = request.POST.get('backdated_at', '').strip()
        if backdated_raw:
            try:
                from datetime import datetime as _dt
                naive = _dt.strptime(backdated_raw, '%Y-%m-%dT%H:%M')
                from django.utils import timezone as _tz
                backdated_at = _tz.make_aware(naive, _tz.get_current_timezone())
            except Exception:
                backdated_at = None

        expiry_date = None
        if trans_type == "Receipt":
            expiry_raw = request.POST.get("expiry_date", "").strip()
            if expiry_raw:
                try:
                    from datetime import date as _date
                    expiry_date = _date.fromisoformat(expiry_raw)
                except (ValueError, TypeError):
                    expiry_date = None

        transaction = Transaction.objects.create(
            item=item,
            type=trans_type,
            qty=quantity,
            recipient=recipient,
            invoice_no=invoice_no,
            business=user_profile.business,
            payment_method=(
                request.POST.get("payment_method", "cash")
                if trans_type == "Issue"
                else ""
            ),
            expiry_date=expiry_date,
            **({"created_at": backdated_at} if backdated_at else {}),
        )

        # ── RESTOCK REQUEST AUTO-RESOLVE ─────────────────────────────────
        if trans_type == 'Receipt':
            _pending_srs = list(
                StockRequest.objects.filter(
                    business=user_profile.business,
                    item=item,
                    status__in=[StockRequest.STATUS_PENDING, StockRequest.STATUS_ORDERED],
                ).order_by('requested_at')
            )
            for _sr in _pending_srs:
                _sr.status = StockRequest.STATUS_RECEIVED
                _sr.received_at = timezone.now()
                _sr.received_by = request.user
                _sr.received_qty = abs(quantity)
                _sr.resolved_txn = transaction
                _sr.save(update_fields=['status', 'received_at', 'received_by', 'received_qty', 'resolved_txn'])
            if _pending_srs:
                restock_resolved = True
                _staff_name_r = request.user.get_full_name() or request.user.username
                _sms_r = (
                    f"✅ {_staff_name_r} amepokea: {abs(quantity)} {item.unit} ya "
                    f"{item.description}. Akiba sasa: {item.current_balance()}."
                )
                for _op in user_profile.business.users.filter(role='owner'):
                    try:
                        Notification.objects.create(
                            user=_op.user,
                            title=f"Stock Received: {item.description}",
                            message=_sms_r,
                            notification_type='info',
                        )
                    except Exception:
                        pass
                    _owner_phone_r = getattr(_op, 'phone', '') or user_profile.business.phone or ''
                    if _owner_phone_r:
                        try:
                            from core.notifications import send_sms_notification, normalize_ke_phone
                            send_sms_notification(_sms_r, normalize_ke_phone(_owner_phone_r))
                        except Exception as _exc_r:
                            logging.getLogger(__name__).error('Restock received SMS failed: %s', _exc_r)
        # ─────────────────────────────────────────────────────────────────

        # ── COST PRICE UPDATE (Receipt only) ──────────────────────────────
        # When receiving stock, the delivered price may differ from the stored
        # cost price. A delivery fee can also be entered — in that case we
        # calculate the landed cost per unit and use that instead.
        #
        # Landed cost per unit = (qty × unit_price + delivery_fee) / qty
        #
        # If the owner ticked "update cost price", Item.cost_price is updated
        # to the landed cost per unit (or just the unit price if no delivery fee).
        if trans_type == "Receipt" and (user_profile.is_owner or getattr(user_profile, 'can_input_cost_price', False)):
            new_cost_price_raw = request.POST.get("new_cost_price", "").strip()
            delivery_fee_raw = request.POST.get("delivery_fee", "0").strip()
            update_cost_price = request.POST.get("update_cost_price") == "on"

            if update_cost_price and new_cost_price_raw:
                try:
                    unit_price = Decimal(new_cost_price_raw)
                    delivery_fee = (
                        Decimal(delivery_fee_raw) if delivery_fee_raw else Decimal("0")
                    )
                    qty_decimal = Decimal(
                        str(abs(quantity))
                    )  # quantity is already negative for issues

                    # Calculate landed cost per unit if there is a delivery fee
                    if delivery_fee > 0 and qty_decimal > 0:
                        landed_cost = (
                            (unit_price * qty_decimal) + delivery_fee
                        ) / qty_decimal
                    else:
                        landed_cost = unit_price

                    old_cost = item.cost_price

                    if landed_cost != old_cost:
                        item.cost_price = landed_cost
                        item.save(update_fields=["cost_price"])

                        if delivery_fee > 0:
                            messages.info(
                                request,
                                _(
                                    "Cost price for %(item)s updated to landed cost KES %(new)s/%(unit)s "
                                    "(KES %(unit_price)s unit + KES %(fee)s delivery ÷ %(qty)s %(unit)s)."
                                )
                                % {
                                    "item": item.description,
                                    "new": f"{landed_cost:,.2f}",
                                    "unit": item.unit,
                                    "unit_price": f"{unit_price:,.2f}",
                                    "fee": f"{delivery_fee:,.2f}",
                                    "qty": f"{qty_decimal:,.0f}",
                                },
                            )
                        else:
                            messages.info(
                                request,
                                _(
                                    "Cost price for %(item)s updated from KES %(old)s to KES %(new)s."
                                )
                                % {
                                    "item": item.description,
                                    "old": f"{old_cost:,.2f}" if old_cost else "—",
                                    "new": f"{landed_cost:,.2f}",
                                },
                            )
                except Exception:
                    pass  # Never block transaction recording due to cost price update failure
        # ─────────────────────────────────────────────────────────────────

        # Notify owner when staff records a receipt (cost price may need updating)
        if trans_type == 'Receipt' and not user_profile.is_owner_or_manager:
            staff_name = request.user.get_full_name() or request.user.username
            cost_str = f'KES {item.cost_price}' if item.cost_price else 'not set'
            notif_msg = (
                f'{staff_name} received {abs(quantity)} {item.unit} of {item.description}. '
                f'Current cost price: {cost_str}. Update in Manage Items if needed.'
            )
            sms_msg = (
                f'COST PRICE NEEDED: {staff_name} received {abs(quantity)} {item.unit} '
                f'of {item.description}. Cost price is {cost_str}. '
                f'Log in to Duka Mwecheche to update it.'
            )
            email_subject = f'Cost price check needed — {item.description} | Duka Mwecheche'
            email_html = f"""
    <div style="font-family:Arial,sans-serif;max-width:500px;margin:0 auto;">
        <h2 style="color:#c9a84c;">📦 Receipt Recorded — Cost Price Needed</h2>
        <p><strong>{staff_name}</strong> has received stock:</p>
        <div style="background:#f5f5f5;padding:1rem;border-radius:8px;margin:1rem 0;">
            <strong style="font-size:1.1rem;">{item.description}</strong><br>
            Quantity received: {abs(quantity)} {item.unit}<br>
            Current cost price: {cost_str}<br>
            Store: {item.store.name if item.store else 'N/A'}
        </div>
        <p>Please <a href="https://www.dukamwecheche.co.ke/manage/items/">update the cost price</a>
        to keep your profit margin calculations accurate.</p>
        <p style="color:#888;font-size:0.85rem;">— Duka Mwecheche</p>
    </div>
    """

            owner_profiles = user_profile.business.users.filter(role='owner')
            for op in owner_profiles:
                # In-app notification
                try:
                    Notification.objects.create(
                        user=op.user,
                        title='Receipt recorded — cost price check needed',
                        message=notif_msg,
                        notification_type='info',
                    )
                except Exception:
                    pass

                # SMS — suppressed when a restock request was resolved (owner already got stock-received SMS)
                if not restock_resolved:
                    try:
                        from core.notifications import send_sms_notification, normalize_ke_phone
                        owner_phone = getattr(op, 'phone', '') or user_profile.business.phone or ''
                        if owner_phone:
                            phone = normalize_ke_phone(owner_phone)
                            if phone:
                                send_sms_notification(sms_msg, phone)
                    except Exception as e:
                        import logging
                        logging.getLogger(__name__).error('Cost price SMS failed: %s', e)

                # Email
                try:
                    from core.notifications import send_email_notification
                    if op.user.email:
                        send_email_notification(op.user.email, email_subject, email_html)
                except Exception as e:
                    import logging
                    logging.getLogger(__name__).error('Cost price email failed: %s', e)

        # ── YIELD: auto-create Wastage transaction for yield items ────────
        # When receiving a yield item (e.g. whole goat → sellable cuts),
        # the usable portion is qty × yield_factor. The remainder is wastage
        # and must be deducted from stock so the balance reflects usable qty.
        if trans_type == 'Receipt' and item.is_yield_item and item.yield_factor:
            received_qty = abs(quantity)  # quantity is positive for receipts
            wastage_qty = Decimal(str(round(float(received_qty) * (1 - float(item.yield_factor)), 4)))
            if wastage_qty > 0:
                Transaction.objects.create(
                    item=item,
                    type='Wastage',
                    qty=-wastage_qty,  # negative to reduce stock
                    recipient='',
                    invoice_no=invoice_no,
                    business=user_profile.business,
                )
                usable_qty = received_qty - wastage_qty
                messages.info(
                    request,
                    _(
                        'Yield applied: %(usable)s %(unit)s usable, '
                        '%(wastage)s %(unit)s wastage recorded (%(pct)s%% yield).'
                    ) % {
                        'usable': float(usable_qty),
                        'unit': item.unit,
                        'wastage': float(wastage_qty),
                        'pct': int(float(item.yield_factor) * 100),
                    },
                )
        # ─────────────────────────────────────────────────────────────────

        # Count today's transactions for SMS/WhatsApp decision
        daily_count = Transaction.objects.filter(
            business=user_profile.business, date=date.today()
        ).count()

        # Send notifications in a background thread — never block the HTTP response
        try:
            from .notifications import notify_transaction_async

            notify_transaction_async(
                transaction.id,
                user_profile.business.id,
                daily_count,
                user_id=request.user.id,
            )
        except Exception:
            pass

        messages.success(
            request,
            _(
                "%(quantity)s %(unit)s of %(item_description)s recorded as %(transaction_type)s."
            )
            % {
                "quantity": abs(quantity),
                "unit": item.unit,
                "item_description": item.description,
                "transaction_type": trans_type.lower(),
            },
        )
        if is_quick:
            return JsonResponse({
                'ok': True,
                'item': item.description,
                'new_balance': float(item.current_balance()),
                'restock_resolved': restock_resolved,
            })
        return redirect("add_transaction")

    items = Item.objects.filter(store__business=user_profile.business).exclude(
        is_keg=True
    ).exclude(store__is_kitchen=True).order_by("material_no")
    restricted_items_data = {}
    if not user_profile.is_owner_or_manager:
        restricted_qs = Item.objects.filter(
            store__business=user_profile.business,
            is_restricted=True,
        ).values('id', 'restricted_quantity')
        for r in restricted_qs:
            restricted_items_data[r['id']] = r['restricted_quantity']

    is_owner = user_profile.is_owner_or_manager
    can_input_cost = is_owner or getattr(user_profile, 'can_input_cost_price', False)
    context = {
        "items": items,
        "stores": stores,
        "customers": customers,
        "today": timezone.now().strftime("%B %d, %Y"),
        "restricted_item_ids": list(restricted_items_data.keys()),
        "restricted_items_data": restricted_items_data,
        "is_owner": is_owner,
        "can_input_cost_price": can_input_cost,
        "show_cost_price_input_only": (not is_owner) and can_input_cost,
    }
    return render(request, "core/add_transaction.html", context)


@login_required
def transaction_history(request):
    user_profile = get_user_profile(request)
    if not user_profile:
        return redirect("home")

    transactions = (
        Transaction.objects.filter(item__store__business=user_profile.business)
        .select_related("item", "item__store")
        .order_by("-date")
    )

    context = {
        "transactions": transactions,
        "today": timezone.now().strftime("%B %d, %Y"),
    }
    return render(request, "core/transaction_history.html", context)


@login_required
def export_transactions_excel(request):
    user_profile = get_user_profile(request)
    if not user_profile:
        return redirect("home")

    transactions = (
        Transaction.objects.filter(item__store__business=user_profile.business)
        .select_related("item", "item__store")
        .order_by("-date")
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transaction History"
    ws.append(
        [
            "Date",
            "Item",
            "Material No",
            "Store",
            "Type",
            "Qty",
            "Recipient",
            "Invoice No",
            "Payment Method",
        ]
    )

    for t in transactions:
        ws.append(
            [
                str(t.date),
                t.item.description,
                t.item.material_no,
                t.item.store.name,
                t.type,
                t.qty,
                t.recipient or "—",
                t.invoice_no or "—",
                t.get_payment_method_display() if t.payment_method else "Cash",
            ]
        )

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=transaction_history.xlsx"
    wb.save(response)
    return response


# ── ITEM DETAIL ───────────────────────────────────────────────────────────────


@login_required
def item_detail(request, item_id):
    user_profile = get_user_profile(request)
    if not user_profile:
        return redirect("home")

    item = get_object_or_404(Item, id=item_id, store__business=user_profile.business)
    transactions = item.transactions.all().order_by("-date")
    context = {
        "item": item,
        "transactions": transactions,
        "today": timezone.now().strftime("%B %d, %Y"),
    }
    return render(request, "core/item_detail.html", context)


@login_required
@owner_or_manager_required
def create_po_from_item(request, item_id):
    """Quick action: create a draft Purchase Order for the item using recommended qty."""
    user_profile = get_user_profile(request)
    if not user_profile:
        return redirect("home")

    item = get_object_or_404(Item, id=item_id, store__business=user_profile.business)
    try:
        qty = item.recommended_order_qty()
    except Exception:
        qty = 0

    if not qty or qty <= 0:
        messages.info(request, _("No order recommended for this item."))
        return redirect("item_detail", item_id=item.id)

    po = PurchaseOrder.objects.create(
        business=user_profile.business,
        status="draft",
        created_by=request.user,
    )
    PurchaseOrderLine.objects.create(
        po=po,
        item=item,
        quantity_ordered=qty,
        unit_price=item.cost_price or 0,
    )
    messages.success(
        request,
        _("Draft purchase order created with %(qty)s units of %(item)s.")
        % {"qty": qty, "item": item.description},
    )
    return redirect("stock_list")


@login_required
@owner_or_manager_required
def purchase_orders_list(request):
    user_profile = get_user_profile(request)
    if not user_profile:
        return redirect("home")
    pos = (
        PurchaseOrder.objects.filter(business=user_profile.business)
        .order_by("-created_at")
        .prefetch_related("lines__item")
    )
    context = {
        "purchase_orders": pos,
        "today": timezone.now().strftime("%B %d, %Y"),
    }
    return render(request, "core/purchase_order_list.html", context)


@login_required
@owner_or_manager_required
def purchase_order_detail(request, po_id):
    user_profile = get_user_profile(request)
    if not user_profile:
        return redirect("home")
    po = get_object_or_404(PurchaseOrder, id=po_id, business=user_profile.business)
    context = {
        "po": po,
        "today": timezone.now().strftime("%B %d, %Y"),
    }
    return render(request, "core/purchase_order_detail.html", context)


@login_required
@owner_or_manager_required
@require_POST
def cancel_purchase_order(request, po_id):
    """'cancelled' was a valid STATUS_CHOICES value with no view that ever
    set it — the only way to reach it was hand-editing the raw status field
    (closed off by PurchaseOrderForm's restricted choices). Allowed from
    draft/ordered/part_received (a supplier can stop delivering partway
    through a real order) but not from received/cancelled (nothing left to
    cancel; cancelling an already-cancelled PO is a harmless no-op redirect,
    not an error, in case of a double-click)."""
    user_profile = get_user_profile(request)
    if not user_profile:
        return redirect("home")
    po = get_object_or_404(PurchaseOrder, id=po_id, business=user_profile.business)

    if po.status == "cancelled":
        messages.info(request, _("This PO is already cancelled."))
    elif po.status == "received":
        messages.error(request, _("This PO is already fully received — nothing to cancel."))
    else:
        po.status = "cancelled"
        po.save(update_fields=["status"])
        messages.success(request, _("Purchase order cancelled."))
    return redirect("purchase_order_detail", po.id)


@login_required
@owner_or_manager_required
def purchase_order_create(request):
    user_profile = get_user_profile(request)
    if not user_profile:
        return redirect("home")

    if request.method == "POST":
        from core.idempotency import claim_checkout_token
        idem_token = (request.POST.get("idempotency_token") or "").strip()
        if not claim_checkout_token(user_profile.business_id, idem_token):
            messages.info(request, _("Hii tayari imehifadhiwa."))
            return redirect("purchase_orders_list")

        form = PurchaseOrderForm(request.POST)
        temp_po = PurchaseOrder(business=user_profile.business)
        formset = PurchaseOrderLineFormSet(request.POST, instance=temp_po)
        # Restrict the item queryset before validation too (defense in
        # depth alongside the per-line business check below) — matches the
        # fix applied to purchase_order_edit, which lacked both layers.
        for f in formset.forms:
            if "item" in f.fields:
                f.fields["item"].queryset = Item.objects.filter(
                    business=user_profile.business
                )
        if form.is_valid() and formset.is_valid():
            po = form.save(commit=False)
            po.business = user_profile.business
            po.created_by = request.user
            po.save()
            formset.instance = po
            lines = formset.save(commit=False)
            for line in lines:
                if line.item and line.item.business != user_profile.business:
                    continue
                line.save()
            messages.success(request, _("Purchase order created."))
            return redirect("purchase_order_detail", po.id)
    else:
        form = PurchaseOrderForm()
        temp_po = PurchaseOrder(business=user_profile.business)
        formset = PurchaseOrderLineFormSet(instance=temp_po)

    for f in formset.forms:
        if "item" in f.fields:
            f.fields["item"].queryset = Item.objects.filter(
                business=user_profile.business
            )

    context = {
        "form": form,
        "formset": formset,
        "today": timezone.now().strftime("%B %d, %Y"),
    }
    return render(request, "core/purchase_order_form.html", context)


@login_required
def item_recommendation(request, item_id):
    """API: return recommended order qty and related metrics for an item."""
    user_profile = get_user_profile(request)
    if not user_profile:
        return JsonResponse({"error": "unauthorized"}, status=403)

    try:
        item = get_object_or_404(
            Item, id=item_id, store__business=user_profile.business
        )
    except Exception:
        return JsonResponse({"error": "not_found_or_unauthorized"}, status=404)

    try:
        recommended = item.recommended_order_qty()
    except Exception:
        recommended = 0

    data = {
        "recommended_qty": int(recommended or 0),
        "on_order": int(item.on_order() or 0),
        "current_balance": int(item.current_balance() or 0),
        "reorder_point": int(item.reorder_point() or 0),
        "target_stock": int(item.target_stock() or 0),
    }
    return JsonResponse(data)


@login_required
def item_search(request):
    """Simple search endpoint for items. Returns JSON list."""
    user_profile = get_user_profile(request)
    if not user_profile:
        return JsonResponse({"results": []})

    q = request.GET.get("q", "").strip()
    items_qs = Item.objects.filter(business=user_profile.business).exclude(is_keg=True)
    if q:
        items_qs = items_qs.filter(
            Q(material_no__icontains=q) | Q(description__icontains=q)
        )
    items_qs = items_qs.order_by("material_no")[:30]
    results = []
    for it in items_qs:
        results.append(
            {
                "id": it.id,
                "material_no": it.material_no,
                "description": it.description,
                "cost_price": float(it.cost_price) if it.cost_price else None,
                "selling_price": float(it.selling_price) if it.selling_price else None,
                "unit": it.unit,
            }
        )
    return JsonResponse({"results": results})


@login_required
def item_cost_price(request, item_id):
    """
    AJAX endpoint: return cost price and key fields for a single item.
    Used by the Add Transaction form to show the cost price confirmation
    section when transaction type = Receipt.

    GET /core/items/<item_id>/cost-price/
    Returns: { cost_price, description, unit, selling_price }
    """
    user_profile = get_user_profile(request)
    if not user_profile:
        return JsonResponse({"error": "unauthorized"}, status=403)

    item = get_object_or_404(Item, id=item_id, store__business=user_profile.business)

    return JsonResponse(
        {
            "cost_price": float(item.cost_price) if item.cost_price else None,
            "description": item.description,
            "unit": item.unit,
            "selling_price": float(item.selling_price) if item.selling_price else None,
            "is_yield_item": item.is_yield_item,
            "yield_factor": float(item.yield_factor) if item.yield_factor else None,
        }
    )


@login_required
@owner_or_manager_required
def purchase_order_edit(request, po_id):
    user_profile = get_user_profile(request)
    if not user_profile:
        return redirect("home")
    po = get_object_or_404(PurchaseOrder, id=po_id, business=user_profile.business)

    if request.method == "POST":
        from core.idempotency import claim_checkout_token
        idem_token = (request.POST.get("idempotency_token") or "").strip()
        if not claim_checkout_token(user_profile.business_id, idem_token):
            messages.info(request, _("Hii tayari imehifadhiwa."))
            return redirect("purchase_order_detail", po.id)

        form = PurchaseOrderForm(request.POST, instance=po)
        formset = PurchaseOrderLineFormSet(request.POST, instance=po)
        # CRITICAL: restrict the item queryset BEFORE validation, not just
        # when re-rendering. purchase_order_create() has its own explicit
        # cross-tenant guard after formset.save(commit=False); this view
        # called formset.save() directly with an unrestricted 'item' field
        # queryset, so ModelChoiceField validation accepted ANY business's
        # item_id and formset.save() persisted it — an authenticated user
        # could inject a PurchaseOrderLine referencing another business's
        # Item, which receive_goods() would then use to create a
        # Transaction against a stranger's Item, corrupting their stock
        # balance (2026-07-21 supply-chain audit finding).
        for f in formset.forms:
            if "item" in f.fields:
                f.fields["item"].queryset = Item.objects.filter(
                    business=user_profile.business
                )
        if form.is_valid() and formset.is_valid():
            form.save()
            lines = formset.save()
            messages.success(request, _("Purchase order updated."))
            return redirect("purchase_order_detail", po.id)
    else:
        form = PurchaseOrderForm(instance=po)
        formset = PurchaseOrderLineFormSet(instance=po)

    for f in formset.forms:
        if "item" in f.fields:
            f.fields["item"].queryset = Item.objects.filter(
                business=user_profile.business
            )

    context = {
        "form": form,
        "formset": formset,
        "po": po,
        "today": timezone.now().strftime("%B %d, %Y"),
    }
    return render(request, "core/purchase_order_form.html", context)


# ── GOODS RECEIPTS ────────────────────────────────────────────────────────────


@login_required
@owner_or_manager_required
def receive_goods(request, po_id):
    """
    Record a delivery against a PurchaseOrder.

    GET  — shows one form row per outstanding PO line, pre-filled with
           the PO unit price as the 'actual price' (edit if delivery price differs).
    POST — saves the GoodsReceipt header + lines, auto-creates Transaction records,
           optionally updates Item cost prices, then updates PO status.
    """
    user_profile = get_user_profile(request)
    if not user_profile:
        return redirect("home")

    po = get_object_or_404(PurchaseOrder, id=po_id, business=user_profile.business)

    if po.status in ("received", "cancelled"):
        messages.error(request, _("This PO is already fully received or cancelled."))
        return redirect("purchase_order_detail", po.id)

    outstanding_lines = [
        line
        for line in po.lines.select_related("item").all()
        if line.quantity_remaining() > 0
    ]

    if not outstanding_lines:
        messages.info(request, _("All items on this PO have already been received."))
        return redirect("purchase_order_detail", po.id)

    if request.method == "POST":
        # Server-side double-submit backstop — see core/idempotency.py. A
        # double-click or slow-network retry on this form would otherwise
        # double-count the physical delivery: two GoodsReceiptLines, two
        # increments of quantity_received, two stock-in Transactions.
        from core.idempotency import claim_checkout_token
        idem_token = (request.POST.get("idempotency_token") or "").strip()
        if not claim_checkout_token(user_profile.business_id, idem_token):
            messages.info(request, _("Hii tayari imehifadhiwa."))
            return redirect("purchase_order_detail", po.id)

        receipt_form = GoodsReceiptForm(request.POST)
        line_formset = GoodsReceiptLineFormSet(request.POST)

        if receipt_form.is_valid() and line_formset.is_valid():
            from django.db import transaction as db_transaction

            with db_transaction.atomic():
                # Re-fetch and lock the PO inside the transaction — the
                # status check above ran before this request took the lock,
                # so two near-simultaneous submissions could otherwise both
                # pass it and both write a full receipt for the same
                # delivery. Re-check status under the lock and bail out if
                # another request already finished receiving this PO.
                locked_po = PurchaseOrder.objects.select_for_update().get(id=po.id)
                if locked_po.status in ("received", "cancelled"):
                    messages.info(
                        request, _("This PO was already received (by another request).")
                    )
                    return redirect("purchase_order_detail", po.id)

                receipt = receipt_form.save(commit=False)
                receipt.po = locked_po
                receipt.received_by = request.user
                receipt.save()

                items_received = 0
                for form in line_formset.forms:
                    data = form.cleaned_data
                    qty = data.get("quantity_received", 0)
                    if not qty:
                        continue

                    po_line = get_object_or_404(
                        PurchaseOrderLine.objects.select_for_update(),
                        id=data["po_line_id"], po=locked_po,
                    )

                    qty = min(qty, po_line.quantity_remaining())
                    if qty <= 0:
                        continue

                    GoodsReceiptLine.objects.create(
                        receipt=receipt,
                        po_line=po_line,
                        quantity_received=qty,
                        actual_unit_price=data["actual_unit_price"],
                        update_cost_price=data.get("update_cost_price", False),
                        notes=data.get("notes", ""),
                    )

                    po_line.quantity_received += qty
                    po_line.save()

                    if data.get("update_cost_price") and data.get("actual_unit_price"):
                        item = po_line.item
                        item.cost_price = data["actual_unit_price"]
                        item.save()

                    invoice_ref = receipt.delivery_note_no or f"GR-{receipt.id}"
                    Transaction.objects.create(
                        business=locked_po.business,
                        item=po_line.item,
                        date=receipt.received_date,
                        invoice_no=invoice_ref,
                        type="Receipt",
                        qty=qty,
                        recipient=f"PO-{locked_po.id}",
                    )
                    items_received += 1

                locked_po.refresh_from_db()
                all_lines = list(locked_po.lines.all())
                if all_lines and all(ln.quantity_remaining() == 0 for ln in all_lines):
                    locked_po.status = "received"
                else:
                    locked_po.status = "part_received"
                locked_po.save()

            if items_received:
                messages.success(
                    request, _("Goods receipt recorded. Stock balances updated.")
                )
            else:
                messages.warning(
                    request, _("No quantities entered — nothing was saved.")
                )
            return redirect("purchase_order_detail", po.id)

    else:
        receipt_form = GoodsReceiptForm(
            initial={"received_date": timezone.now().date()}
        )
        initial_data = [
            {
                "po_line_id": line.id,
                "quantity_received": line.quantity_remaining(),
                "actual_unit_price": line.unit_price or 0,
                "update_cost_price": False,
            }
            for line in outstanding_lines
        ]
        line_formset = GoodsReceiptLineFormSet(initial=initial_data)

    line_form_pairs = list(zip(outstanding_lines, line_formset.forms))

    context = {
        "po": po,
        "receipt_form": receipt_form,
        "line_formset": line_formset,
        "line_form_pairs": line_form_pairs,
        "today": timezone.now().strftime("%B %d, %Y"),
    }
    return render(request, "core/receive_goods.html", context)


@login_required
@owner_or_manager_required
def goods_receipt_detail(request, receipt_id):
    """View a single goods receipt with line-by-line variance analysis."""
    user_profile = get_user_profile(request)
    if not user_profile:
        return redirect("home")

    receipt = get_object_or_404(
        GoodsReceipt.objects.select_related("po", "received_by").prefetch_related(
            "lines__po_line__item"
        ),
        id=receipt_id,
        po__business=user_profile.business,
    )
    context = {
        "receipt": receipt,
        "today": timezone.now().strftime("%B %d, %Y"),
    }
    return render(request, "core/goods_receipt_detail.html", context)


# ── EXPORT STOCK ──────────────────────────────────────────────────────────────


@login_required
def export_stock_excel(request):
    user_profile = get_user_profile(request)
    store_id = request.GET.get("store")

    if user_profile:
        items = Item.objects.filter(store__business=user_profile.business).exclude(store__is_kitchen=True)
        if store_id:
            items = items.filter(store_id=store_id)
    else:
        items = Item.objects.none()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stock List"
    ws.append(
        [
            "Material No",
            "Description",
            "Unit",
            "Current Balance",
            "Reorder Level",
            "Selling Price",
            "Status",
            "Store",
        ]
    )

    for item in items:
        status = (
            "OUT OF STOCK"
            if item.current_balance() <= 0
            else "REORDER" if item.needs_reorder() else "AVAILABLE"
        )
        ws.append(
            [
                item.material_no,
                item.description,
                item.unit,
                item.current_balance(),
                item.reorder_level,
                float(item.selling_price) if item.selling_price else "",
                status,
                item.store.name,
            ]
        )

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=stock_list.xlsx"
    wb.save(response)
    return response


# ── MANAGE ITEMS ──────────────────────────────────────────────────────────────


@login_required
@owner_or_manager_required
def manage_items(request):
    user_profile = request.user.userprofile
    items = (
        Item.objects.filter(business=user_profile.business)
        .select_related("store")
        .order_by("store__name", "description")
    )

    context = {
        "items": items,
        "today": timezone.now().strftime("%B %d, %Y"),
    }
    return render(request, "core/manage_items.html", context)


def _resolve_category(cat_text):
    """Resolve a free-text category name to a Category, creating one if it doesn't exist.

    Fixes the bug where a typed category (e.g. "Vegetables") was silently dropped
    when it didn't already match an existing Category record.
    """
    from django.utils.text import slugify
    cat_text = (cat_text or '').strip()
    if not cat_text:
        return None
    cat = Category.objects.filter(level1__iexact=cat_text).first()
    if cat:
        return cat
    base = (slugify(cat_text) or 'category')[:40]
    code, n = base, 1
    while Category.objects.filter(code=code).exists():
        n += 1
        code = f"{base}-{n}"[:50]
    return Category.objects.create(code=code, level1=cat_text)


@login_required
@owner_or_manager_required
def add_item(request):
    user_profile = request.user.userprofile

    if request.method == "POST":
        form = ItemForm(
            request.POST, business=user_profile.business, show_cost_price=True
        )
        if form.is_valid():
            item = form.save(commit=False)
            item.business = user_profile.business
            if not item.material_no:
                last_item = (
                    Item.objects.filter(business=user_profile.business)
                    .order_by("id")
                    .last()
                )
                next_id = (last_item.id + 1) if last_item else 1
                item.material_no = f"MAT-{next_id:04d}"
            # Handle yield_factor from percentage input
            if request.POST.get('is_yield_item') == 'on':
                item.is_yield_item = True
                try:
                    yield_pct = Decimal(request.POST.get('yield_percentage', '0'))
                    item.yield_factor = yield_pct / Decimal('100')
                except (InvalidOperation, ValueError):
                    item.yield_factor = Decimal('1.0')
            else:
                item.is_yield_item = False
                item.yield_factor = None
            # Resolve / create category from the free-text input
            cat_obj = _resolve_category(request.POST.get('cat_text_input'))
            if cat_obj:
                item.category = cat_obj
            item.save()
            # Handle owner/manager fields (restrictions + produce)
            if user_profile.is_owner_or_manager:
                item.is_restricted = request.POST.get('is_restricted') == 'on'
                item.restriction_notes = request.POST.get('restriction_notes', '').strip()
                try:
                    item.restricted_quantity = max(0, int(request.POST.get('restricted_quantity', 0)))
                except (ValueError, TypeError):
                    item.restricted_quantity = 0
                item.is_kitchen_batch = request.POST.get('is_kitchen_batch') == 'on'
                item.is_produce = request.POST.get('is_produce') == 'on' and not item.is_kitchen_batch
                pmode = request.POST.get('produce_mode', 'PORTION')
                item.produce_mode = pmode if pmode in ('PORTION', 'BUNCH') else 'PORTION'
                # km_mix_group is submitted by kitchen batch items via a hidden input to avoid
                # conflict with the produce section's mix_group_input (which stays empty in kitchen mode).
                item.mix_group = (request.POST.get('km_mix_group') or request.POST.get('mix_group') or '').strip()[:40]
                try:
                    # km_revenue_multiplier from kitchen batch form takes priority
                    rm_raw = request.POST.get('km_revenue_multiplier') or request.POST.get('revenue_multiplier') or '1.70'
                    rm = Decimal(rm_raw)
                    item.revenue_multiplier = rm if rm > 0 else Decimal('1.70')
                except (ValueError, InvalidOperation):
                    item.revenue_multiplier = Decimal('1.70')
                item.is_keg = request.POST.get('is_keg') == 'on'
                item.bottle_envelope = request.POST.get('bottle_envelope') == 'on'
                try:
                    item.volume_ml = int(request.POST.get('volume_ml')) if request.POST.get('volume_ml') else None
                except (ValueError, TypeError):
                    item.volume_ml = None
                try:
                    item.tot_ml = Decimal(request.POST.get('tot_ml')) if request.POST.get('tot_ml') else None
                except InvalidOperation:
                    item.tot_ml = None
                try:
                    item.tots_per_unit = Decimal(request.POST.get('tots_per_unit')) if request.POST.get('tots_per_unit') else None
                except InvalidOperation:
                    item.tots_per_unit = None
                item.save(update_fields=['is_restricted', 'restriction_notes', 'restricted_quantity',
                                         'is_kitchen_batch', 'is_produce', 'produce_mode', 'mix_group',
                                         'revenue_multiplier', 'is_keg', 'bottle_envelope',
                                         'volume_ml', 'tot_ml', 'tots_per_unit'])

                # ── PRODUCE / KITCHEN BATCH PORTION PRESETS ──────────────────
                preset_labels   = request.POST.getlist('preset_label')
                preset_prices   = request.POST.getlist('preset_price')
                preset_qty      = request.POST.getlist('preset_qty_consumed')
                preset_ids      = request.POST.getlist('preset_id')
                preset_servings = request.POST.getlist('preset_serving_type')
                preset_khakis   = request.POST.getlist('preset_khaki_type')

                submitted_ids = [int(pid) for pid in preset_ids if pid.strip()]
                item.portion_presets.exclude(id__in=submitted_ids).delete()

                is_bunch = item.produce_mode == 'BUNCH'
                for i, label in enumerate(preset_labels):
                    label = label.strip()
                    try:
                        price = Decimal(preset_prices[i])
                    except (ValueError, InvalidOperation, IndexError):
                        continue
                    if not label:
                        if is_bunch or item.is_kitchen_batch:
                            label = f"Ya {price:g}"
                        else:
                            continue
                    try:
                        qty_c = (Decimal(preset_qty[i])
                                 if i < len(preset_qty) and str(preset_qty[i]).strip()
                                 else Decimal('0'))
                    except (ValueError, InvalidOperation, IndexError):
                        qty_c = Decimal('0')
                    order = i
                    serving = preset_servings[i] if i < len(preset_servings) else 'cup'
                    if serving not in ('cup', 'pint', 'jug'):
                        serving = 'cup'
                    khaki = preset_khakis[i] if i < len(preset_khakis) else 'NONE'
                    if khaki not in ('NONE', 'SMALL', 'LARGE'):
                        khaki = 'NONE'
                    pid = preset_ids[i].strip() if i < len(preset_ids) else ''
                    if pid:
                        ItemPortionPreset.objects.filter(id=int(pid), item=item).update(
                            label=label, price=price, quantity_consumed=qty_c,
                            display_order=order, serving_type=serving, khaki_type=khaki,
                        )
                    else:
                        ItemPortionPreset.objects.create(
                            item=item, label=label, price=price,
                            quantity_consumed=qty_c, display_order=order,
                            serving_type=serving, khaki_type=khaki,
                        )
                # ─────────────────────────────────────────────────────────────

            messages.success(
                request,
                _("'%(item_description)s' added successfully.")
                % {"item_description": item.description},
            )
            return redirect("manage_items")
    else:
        form = ItemForm(business=user_profile.business, show_cost_price=True)

    from .business_profiles import get_profile as _get_profile
    import json as _json
    _catalog = _get_profile(user_profile.business).get('catalog', [])
    _kitchen_store_ids = list(
        Store.objects.filter(business=user_profile.business, is_kitchen=True)
        .values_list('id', flat=True)
    )
    context = {
        "form": form,
        "today": timezone.now().strftime("%B %d, %Y"),
        "action": _("Add"),
        "is_add": True,
        "catalog_json": _json.dumps(_catalog),
        "kitchen_store_ids_json": _json.dumps(_kitchen_store_ids),
    }
    return render(request, "core/item_form.html", context)


@login_required
@owner_or_manager_required
def edit_item(request, item_id):
    user_profile = request.user.userprofile
    item = get_object_or_404(Item, id=item_id, business=user_profile.business)

    if request.method == "POST":
        form = ItemForm(
            request.POST,
            instance=item,
            business=user_profile.business,
            show_cost_price=True,
        )
        if form.is_valid():
            item = form.save()
            # Resolve / create category from the free-text input
            cat_obj = _resolve_category(request.POST.get('cat_text_input'))
            if cat_obj:
                item.category = cat_obj
                item.save(update_fields=['category'])
            # Handle yield_factor from percentage input
            if request.POST.get('is_yield_item') == 'on':
                item.is_yield_item = True
                try:
                    yield_pct = Decimal(request.POST.get('yield_percentage', '0'))
                    item.yield_factor = yield_pct / Decimal('100')
                except (InvalidOperation, ValueError):
                    item.yield_factor = Decimal('1.0')
            else:
                item.is_yield_item = False
                item.yield_factor = None
            item.save(update_fields=['is_yield_item', 'yield_factor'])
            # Handle owner/manager fields (restrictions + produce)
            if user_profile.is_owner_or_manager:
                item.is_restricted = request.POST.get('is_restricted') == 'on'
                item.restriction_notes = request.POST.get('restriction_notes', '').strip()
                try:
                    item.restricted_quantity = max(0, int(request.POST.get('restricted_quantity', 0)))
                except (ValueError, TypeError):
                    item.restricted_quantity = 0
                item.is_kitchen_batch = request.POST.get('is_kitchen_batch') == 'on'
                item.is_produce = request.POST.get('is_produce') == 'on' and not item.is_kitchen_batch
                pmode = request.POST.get('produce_mode', 'PORTION')
                item.produce_mode = pmode if pmode in ('PORTION', 'BUNCH') else 'PORTION'
                # km_mix_group is submitted by kitchen batch items via a hidden input to avoid
                # conflict with the produce section's mix_group_input (which stays empty in kitchen mode).
                item.mix_group = (request.POST.get('km_mix_group') or request.POST.get('mix_group') or '').strip()[:40]
                try:
                    # km_revenue_multiplier from kitchen batch form takes priority
                    rm_raw = request.POST.get('km_revenue_multiplier') or request.POST.get('revenue_multiplier') or '1.70'
                    rm = Decimal(rm_raw)
                    item.revenue_multiplier = rm if rm > 0 else Decimal('1.70')
                except (ValueError, InvalidOperation):
                    item.revenue_multiplier = Decimal('1.70')
                item.is_keg = request.POST.get('is_keg') == 'on'
                item.bottle_envelope = request.POST.get('bottle_envelope') == 'on'
                try:
                    item.volume_ml = int(request.POST.get('volume_ml')) if request.POST.get('volume_ml') else None
                except (ValueError, TypeError):
                    item.volume_ml = None
                try:
                    item.tot_ml = Decimal(request.POST.get('tot_ml')) if request.POST.get('tot_ml') else None
                except InvalidOperation:
                    item.tot_ml = None
                try:
                    item.tots_per_unit = Decimal(request.POST.get('tots_per_unit')) if request.POST.get('tots_per_unit') else None
                except InvalidOperation:
                    item.tots_per_unit = None
                item.save(update_fields=['is_restricted', 'restriction_notes', 'restricted_quantity',
                                         'is_kitchen_batch', 'is_produce', 'produce_mode', 'mix_group',
                                         'revenue_multiplier', 'is_keg', 'bottle_envelope',
                                         'volume_ml', 'tot_ml', 'tots_per_unit'])

                # ── PRODUCE / KITCHEN BATCH PORTION PRESETS ──────────────────
                preset_labels   = request.POST.getlist('preset_label')
                preset_prices   = request.POST.getlist('preset_price')
                preset_qty      = request.POST.getlist('preset_qty_consumed')
                preset_ids      = request.POST.getlist('preset_id')
                preset_servings = request.POST.getlist('preset_serving_type')
                preset_khakis   = request.POST.getlist('preset_khaki_type')

                submitted_ids = [int(pid) for pid in preset_ids if pid.strip()]
                item.portion_presets.exclude(id__in=submitted_ids).delete()

                is_bunch = item.produce_mode == 'BUNCH'
                for i, label in enumerate(preset_labels):
                    label = label.strip()
                    try:
                        price = Decimal(preset_prices[i])
                    except (ValueError, InvalidOperation, IndexError):
                        continue
                    if not label:
                        if is_bunch or item.is_kitchen_batch:
                            label = f"Ya {price:g}"
                        else:
                            continue
                    try:
                        qty_c = (Decimal(preset_qty[i])
                                 if i < len(preset_qty) and str(preset_qty[i]).strip()
                                 else Decimal('0'))
                    except (ValueError, InvalidOperation, IndexError):
                        qty_c = Decimal('0')
                    order = i
                    serving = preset_servings[i] if i < len(preset_servings) else 'cup'
                    if serving not in ('cup', 'pint', 'jug'):
                        serving = 'cup'
                    khaki = preset_khakis[i] if i < len(preset_khakis) else 'NONE'
                    if khaki not in ('NONE', 'SMALL', 'LARGE'):
                        khaki = 'NONE'
                    pid = preset_ids[i].strip() if i < len(preset_ids) else ''
                    if pid:
                        ItemPortionPreset.objects.filter(id=int(pid), item=item).update(
                            label=label, price=price, quantity_consumed=qty_c,
                            display_order=order, serving_type=serving, khaki_type=khaki,
                        )
                    else:
                        ItemPortionPreset.objects.create(
                            item=item, label=label, price=price,
                            quantity_consumed=qty_c, display_order=order,
                            serving_type=serving, khaki_type=khaki,
                        )
                # ─────────────────────────────────────────────────────────────

            messages.success(
                request,
                _("'%(item_description)s' updated successfully.")
                % {"item_description": item.description},
            )
            return redirect("manage_items")
    else:
        form = ItemForm(
            instance=item, business=user_profile.business, show_cost_price=True
        )

    from .business_profiles import get_profile as _get_profile
    import json as _json
    _catalog = _get_profile(user_profile.business).get('catalog', [])
    _kitchen_store_ids = list(
        Store.objects.filter(business=user_profile.business, is_kitchen=True)
        .values_list('id', flat=True)
    )
    context = {
        "form": form,
        "item": item,
        "today": timezone.now().strftime("%B %d, %Y"),
        "action": _("Edit"),
        "is_add": False,
        "catalog_json": _json.dumps(_catalog),
        "kitchen_store_ids_json": _json.dumps(_kitchen_store_ids),
    }
    return render(request, "core/item_form.html", context)


@login_required
@owner_or_manager_required
def delete_item(request, item_id):
    user_profile = request.user.userprofile
    item = get_object_or_404(Item, id=item_id, business=user_profile.business)

    if request.method == "POST":
        item_name = item.description
        item.delete()
        messages.success(
            request,
            _("'%(item_name)s' deleted successfully.") % {"item_name": item_name},
        )
        return redirect("manage_items")

    context = {
        "item": item,
        "today": timezone.now().strftime("%B %d, %Y"),
    }
    return render(request, "core/delete_item.html", context)


# ── MANAGE STORES ─────────────────────────────────────────────────────────────


@login_required
@owner_or_manager_required
def manage_stores(request):
    user_profile = request.user.userprofile
    stores = Store.objects.filter(business=user_profile.business)

    if request.method == "POST":
        store_name = request.POST.get("name", "").strip()
        if store_name:
            Store.objects.create(name=store_name, business=user_profile.business)
            messages.success(
                request,
                _("Store '%(store_name)s' created successfully.")
                % {"store_name": store_name},
            )
            return redirect("manage_stores")
        else:
            messages.error(request, _("Store name cannot be empty."))

    context = {
        "stores": stores,
        "today": timezone.now().strftime("%B %d, %Y"),
    }
    return render(request, "core/manage_stores.html", context)


# ── CUSTOMERS ─────────────────────────────────────────────────────────────────


@login_required
def customer_list(request):
    user_profile = get_user_profile(request)
    if not user_profile:
        return redirect("home")

    from core.models import County as CoreCounty
    customers = Customer.objects.filter(business=user_profile.business).select_related('county')
    context = {
        "customers": customers,
        "counties": CoreCounty.objects.all().order_by('name'),
    }
    return render(request, "core/customer_list.html", context)


@login_required
def add_customer(request):
    user_profile = get_user_profile(request)
    if not user_profile:
        return redirect("home")

    from core.models import County as CoreCounty

    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        phone = request.POST.get("phone", "").strip()
        location = request.POST.get("location", "").strip()
        county_id = request.POST.get("county_id", "").strip()

        if name:
            county = None
            if county_id:
                try:
                    county = CoreCounty.objects.get(id=county_id)
                except CoreCounty.DoesNotExist:
                    pass
            Customer.objects.create(
                business=user_profile.business,
                name=name,
                phone=phone,
                location=location,
                county=county,
            )
            messages.success(
                request,
                _("Customer '%(customer_name)s' added.") % {"customer_name": name},
            )
            return redirect("customer_list")
        else:
            messages.error(request, _("Customer name is required."))

    return render(
        request,
        "core/customer_list.html",
        {
            "customers": Customer.objects.filter(business=user_profile.business).select_related('county'),
            "counties": CoreCounty.objects.all().order_by('name'),
        },
    )


@login_required
@owner_or_manager_required
def delete_customer(request, customer_id):
    user_profile = request.user.userprofile
    customer = get_object_or_404(
        Customer, id=customer_id, business=user_profile.business
    )

    if request.method == "POST":
        customer.delete()
        messages.success(request, _("Customer deleted."))
    return redirect("customer_list")


# ── AJAX ──────────────────────────────────────────────────────────────────────


@login_required
def ajax_customers(request):
    """Returns customers for the transaction form dropdown."""
    user_profile = get_user_profile(request)
    if not user_profile:
        return JsonResponse([], safe=False)

    customers = Customer.objects.filter(business=user_profile.business).values(
        "id", "name", "phone"
    )
    return JsonResponse(list(customers), safe=False)


# ── SALES & P&L ───────────────────────────────────────────────────────────────


def get_date_range(period, date_from=None, date_to=None):
    """Returns (start_date, end_date) based on period filter."""
    today = date.today()
    if period == "today":
        return today, today
    elif period == "week":
        return today - timedelta(days=today.weekday()), today
    elif period == "month":
        return today.replace(day=1), today
    elif period == "year":
        return today.replace(month=1, day=1), today
    elif period == "custom" and date_from and date_to:
        try:
            from datetime import datetime

            return (
                datetime.strptime(date_from, "%Y-%m-%d").date(),
                datetime.strptime(date_to, "%Y-%m-%d").date(),
            )
        except ValueError:
            return today.replace(day=1), today
    else:
        return today.replace(day=1), today


@login_required
@owner_or_manager_required
def sales_dashboard(request):
    user_profile = request.user.userprofile
    business = user_profile.business

    period = request.GET.get("period", "month")
    date_from = request.GET.get("date_from")
    date_to = request.GET.get("date_to")
    start_date, end_date = get_date_range(period, date_from, date_to)

    sales = Transaction.objects.filter(
        business=business,
        type="Issue",
        date__gte=start_date,
        date__lte=end_date,
    ).exclude(payment_method='void').select_related("item", "keg_barrel", "produce_bunch")

    def _sd_units(t):
        """Count servings: keg pours and bunch sales count as 1; others use abs(qty)."""
        if getattr(t, 'produce_bunch_id', None) is not None:
            return 1.0
        if getattr(t, 'keg_barrel_id', None) is not None:
            return 1.0
        return float(abs(t.qty or 0))

    total_revenue = sum(t.revenue() for t in sales)
    total_cost = sum(t.cost() for t in sales)
    total_profit = total_revenue - total_cost
    total_units_sold = sum(_sd_units(t) for t in sales)

    all_items = Item.objects.filter(business=business)
    stock_value = sum(item.stock_value() for item in all_items)

    from collections import defaultdict

    daily_revenue = defaultdict(float)
    daily_profit = defaultdict(float)
    for t in sales:
        day_str = str(t.date)
        daily_revenue[day_str] += t.revenue()
        daily_profit[day_str] += t.profit()

    sorted_dates = sorted(daily_revenue.keys())
    chart_labels = sorted_dates
    chart_revenue = [round(daily_revenue[d], 2) for d in sorted_dates]
    chart_profit = [round(daily_profit[d], 2) for d in sorted_dates]

    item_sales = defaultdict(
        lambda: {
            "description": "",
            "units_sold": 0,
            "revenue": 0.0,
            "cost": 0.0,
            "profit": 0.0,
        }
    )
    for t in sales:
        key = t.item.id
        item_sales[key]["description"] = t.item.description
        item_sales[key]["material_no"] = t.item.material_no
        item_sales[key]["units_sold"] += _sd_units(t)
        item_sales[key]["revenue"] += t.revenue()
        item_sales[key]["cost"] += t.cost()
        item_sales[key]["profit"] += t.profit()

    item_sales_list = sorted(
        item_sales.values(), key=lambda x: x["revenue"], reverse=True
    )
    top_items = sorted(item_sales_list, key=lambda x: x["units_sold"], reverse=True)[:5]

    sold_item_ids = set(t.item.id for t in sales)
    slow_items = all_items.exclude(id__in=sold_item_ids)[:10]

    profit_margin = (
        round((total_profit / total_revenue * 100), 1) if total_revenue > 0 else 0
    )

    # Owner Drawings — stock consumed by the owner at cost, shown below gross profit
    owner_txns = Transaction.objects.filter(
        business=business,
        type='OwnerConsumption',
        date__gte=start_date,
        date__lte=end_date,
    ).select_related('item')
    owner_drawings_cost = round(sum(
        abs(float(t.qty or 0)) * float(t.item.cost_price or 0)
        for t in owner_txns
    ), 2)
    net_profit_after_drawings = round(total_profit - owner_drawings_cost, 2)

    context = {
        "period": period,
        "date_from": start_date,
        "date_to": end_date,
        "total_revenue": round(total_revenue, 2),
        "total_cost": round(total_cost, 2),
        "total_profit": round(total_profit, 2),
        "profit_margin": profit_margin,
        "total_units_sold": total_units_sold,
        "stock_value": round(stock_value, 2),
        "owner_drawings_cost": owner_drawings_cost,
        "net_profit_after_drawings": net_profit_after_drawings,
        "item_sales_list": item_sales_list,
        "top_items": top_items,
        "slow_items": slow_items,
        "chart_labels": json.dumps(chart_labels),
        "chart_revenue": json.dumps(chart_revenue),
        "chart_profit": json.dumps(chart_profit),
        "today": timezone.now().strftime("%B %d, %Y"),
    }
    return render(request, "core/sales_dashboard.html", context)


@login_required
@owner_or_manager_required
def export_sales_excel(request):
    user_profile = request.user.userprofile
    business = user_profile.business

    period = request.GET.get("period", "month")
    date_from = request.GET.get("date_from")
    date_to = request.GET.get("date_to")
    start_date, end_date = get_date_range(period, date_from, date_to)

    sales = (
        Transaction.objects.filter(
            business=business,
            type="Issue",
            date__gte=start_date,
            date__lte=end_date,
        )
        .exclude(payment_method='void')
        .select_related("item")
        .order_by("date")
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Report"
    ws.append(
        [
            "Date",
            "Item",
            "Material No",
            "Units Sold",
            "Selling Price",
            "Cost Price",
            "Revenue",
            "Cost",
            "Profit",
            "Recipient",
            "Invoice No",
        ]
    )

    for t in sales:
        ws.append(
            [
                str(t.date),
                t.item.description,
                t.item.material_no,
                abs(t.qty),
                float(t.item.selling_price) if t.item.selling_price else "",
                float(t.item.cost_price) if t.item.cost_price else "",
                round(t.revenue(), 2),
                round(t.cost(), 2),
                round(t.profit(), 2),
                t.recipient or "—",
                t.invoice_no or "—",
            ]
        )

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=sales_report.xlsx"
    wb.save(response)
    return response


# ── NOTIFICATIONS ─────────────────────────────────────────────────────────────


@login_required
def notifications_list(request):
    notifications = request.user.app_notifications.all()[:50]
    request.user.app_notifications.filter(is_read=False).update(is_read=True)
    return render(request, "core/notifications.html", {"notifications": notifications})


def notifications_count(request):
    if not request.user.is_authenticated:
        return JsonResponse({"count": 0, "prompts_count": 0})
    count = request.user.app_notifications.filter(is_read=False).count()
    prompts_count = 0
    profile = getattr(request.user, "userprofile", None)
    if profile and profile.business:
        from .models import PendingTransactionPrompt

        prompts_count = PendingTransactionPrompt.objects.filter(
            business=profile.business, status="pending"
        ).count()
    return JsonResponse({"count": count, "prompts_count": prompts_count})


def daily_summary_webhook(request):
    """Endpoint called by cron-job.org to trigger daily summaries."""
    token = request.GET.get("token")
    expected = os.getenv("CRON_SECRET", "duka-mwecheche-cron-2026")
    if token != expected:
        from django.http import HttpResponseForbidden

        return HttpResponseForbidden("Invalid token")

    from accounts.models import Business
    from .notifications import send_daily_summary

    logger = logging.getLogger(__name__)
    businesses = Business.objects.all().iterator(chunk_size=10)
    count = 0
    for business in businesses:
        count += 1
        try:
            send_daily_summary(business)
        except Exception:
            logger.exception("Daily summary failed for business %s", business.id)

    return JsonResponse({"status": "ok", "businesses": count})


# ── QUICK SELL ────────────────────────────────────────────────────────────────


@login_required
def quick_sell(request):
    user_profile = get_user_profile(request)
    if not user_profile:
        return redirect("home")

    success_data = None

    if request.method == "POST":
        # Shift gate: ALL staff (any business type) must have their own open shift
        if not user_profile.is_owner_or_manager:
            from core.shift_views import get_active_staff_shift
            if get_active_staff_shift(user_profile, user_profile.business) is False:
                messages.error(
                    request,
                    'Fungua shift yako kwanza kabla ya kuuza.'
                )
                return redirect('quick_sell')

        # Server-side double-submit backstop — see core/idempotency.py. Client-side
        # guards (button disable, JS flag) only cover a second click on the same
        # live page; this catches real duplicate requests (slow-network retry,
        # back-button resubmission of the real <form> this view is posted from).
        from core.idempotency import claim_checkout_token
        idem_token = request.POST.get("idempotency_token", "").strip()
        if not claim_checkout_token(user_profile.business_id, idem_token):
            messages.info(request, 'Mauzo haya tayari yamehifadhiwa.')
            return redirect('quick_sell')

        cart_json = request.POST.get("cart", "[]")
        try:
            cart = json.loads(cart_json)
        except (json.JSONDecodeError, TypeError):
            cart = []

        credit_recipient = request.POST.get("recipient", "").strip()
        credit_phone     = request.POST.get("credit_phone", "").strip()
        payment_method_raw = request.POST.get("payment_method", "cash")
        # 'tab' is a UI-only value — persisted to Transaction as 'credit' (same
        # as keg tab convention) but routed to BarTab instead of debt tracker.
        is_tab_sale = (payment_method_raw == "tab")
        payment_method_qs = "credit" if is_tab_sale else payment_method_raw

        # ── CREDIT DISCIPLINE GATE (credit sales only — not bar tabs; tab
        #    creation doesn't use the debt ledger credit_approved path) ────────
        if payment_method_raw == 'credit' and credit_recipient:
            from core.models import Customer as _CustomerModel
            from core.credit_policy import evaluate_credit
            _cust_gate = _CustomerModel.objects.filter(
                business=user_profile.business, name=credit_recipient
            ).first()
            if _cust_gate is None:
                # Auto-create approved — owner is initiating this credit sale
                # right now, which implies approval. credit_approved=False is for
                # when the owner explicitly revokes from the customer profile.
                _cust_gate = _CustomerModel.objects.create(
                    business=user_profile.business,
                    name=credit_recipient,
                    phone=credit_phone,
                    credit_approved=True,
                )
            _decision = evaluate_credit(user_profile.business, _cust_gate)
            if not _decision.allowed:
                messages.error(
                    request,
                    f'Deni haliwezi kutolewa: {_decision.reason} — '
                    'Lipa kwa cash au M-Pesa badala yake.'
                )
                return redirect('quick_sell')
        # ─────────────────────────────────────────────────────────────────────

        recorded = []
        last_transaction = None
        tab_transactions = []  # (transaction, description, amount) for BarTabEntry creation

        for entry in cart:
            # ── Greens / bunch (revenue-envelope) lines ──────────────────
            if entry.get("mode") in ("bunch", "mix"):
                from .produce_views import handle_bunch_cart_entry
                line, txn = handle_bunch_cart_entry(
                    entry,
                    user_profile.business,
                    payment_method_qs,
                    recipient=credit_recipient if payment_method_qs == "credit" else "",
                    recorded_by=request.user,
                )
                if line:
                    recorded.append(line)
                    if txn:
                        last_transaction = txn
                else:
                    # Unlike the regular-item path below, a depleted/closed bunch
                    # used to fail completely silently — the client already
                    # blocks adding an empty bunch tile to the cart, but that
                    # check is against a snapshot fetched when the greens board
                    # last loaded, not at checkout time, so a concurrent sale can
                    # still deplete it in the gap between tap and checkout. If
                    # this was the only line in the cart, the whole request used
                    # to end in total silence — no success, no error, the sale
                    # just vanished (Quick-Sell-module audit finding, 2026-07-19).
                    messages.warning(
                        request,
                        _("Skipped %(name)s: no stock available.")
                        % {"name": entry.get("label") or entry.get("name") or _("item")},
                    )
                continue
            # ─────────────────────────────────────────────────────────────
            item = Item.objects.filter(
                id=entry.get("id"), store__business=user_profile.business
            ).first()
            if not item:
                continue

            # stock_qty = actual stock consumed (may be fractional for produce)
            # display_qty = what to show on receipt (1 for produce portions, integer for normal)
            # For preset entries, stock_qty is per-serving and must be multiplied by cart qty.
            # For regular items, stock_qty is absent and entry["qty"] IS the total qty.
            try:
                raw_stock_qty = Decimal(str(entry.get("stock_qty", entry.get("qty", 0))))
                cart_qty = Decimal(str(entry.get("qty", 1)))
                stock_qty = raw_stock_qty * cart_qty if entry.get("stock_qty") is not None else raw_stock_qty
            except Exception:
                stock_qty = Decimal('0')
            if stock_qty <= 0:
                continue
            display_qty = entry.get("qty", int(stock_qty))
            display_price = float(entry.get("price", 0)) or float(item.selling_price or 0)

            # ── RESTRICTED ITEM CHECK ─────────────────────────────────────
            can_override = getattr(user_profile, 'can_override_restrictions', False)
            if item.is_restricted and not user_profile.is_owner_or_manager and not can_override:
                reserved = item.restricted_quantity or 0
                balance_after = item.current_balance() - stock_qty
                if reserved == 0 or balance_after < reserved:
                    messages.warning(
                        request,
                        _(f'{item.description} requires owner approval for this quantity. '
                          f'Use Add Transaction to submit an approval request.')
                    )
                    continue
                # else: sale is within freely-sellable range — falls through
            # ─────────────────────────────────────────────────────────────

            if item.current_balance() < stock_qty:
                messages.warning(
                    request,
                    _(
                        "Skipped %(item_description)s: only %(available)s %(unit)s in stock."
                    )
                    % {
                        "item_description": item.description,
                        "available": item.current_balance(),
                        "unit": item.unit,
                    },
                )
                continue

            # For portion-preset sales the cart supplies an explicit price that may
            # differ from selling_price × stock_qty (e.g. Tatu mbao: 3 onions for KES 20
            # instead of 3 × KES 10 = KES 30). Store it as sale_amount so revenue() is
            # correct. Only set when stock_qty is explicitly in the cart entry (i.e. a
            # preset tap) — normal item taps omit stock_qty.
            sale_amt = None
            if entry.get("stock_qty") is not None and display_price:
                sale_amt = Decimal(str(round(display_price * float(display_qty), 2)))

            line_amount = Decimal(str(round(display_price * float(display_qty), 2)))
            last_transaction = Transaction.objects.create(
                item=item,
                type="Issue",
                qty=-stock_qty,
                business=user_profile.business,
                payment_method=payment_method_qs,
                sale_amount=sale_amt,
                recipient=credit_recipient if payment_method_qs == "credit" else "",
                recorded_by=request.user,
            )
            recorded.append(
                {
                    "name": item.description,
                    "qty": float(display_qty),
                    "subtotal": float(line_amount),
                }
            )
            if is_tab_sale:
                tab_transactions.append((last_transaction, f"{item.description} ×{display_qty}", line_amount))

        if recorded and last_transaction:
            total = sum(r["subtotal"] for r in recorded)

            try:
                from .notifications import notify_transaction_async

                daily_count = Transaction.objects.filter(
                    business=user_profile.business, date=date.today()
                ).count()
                notify_transaction_async(
                    last_transaction.id,
                    user_profile.business.id,
                    daily_count,
                    user_id=request.user.id,
                )
            except Exception:
                pass

            # ── TAB SALE: create/extend BarTab and attach BarTabEntry records ──
            if is_tab_sale and tab_transactions:
                # Anonymous tab — busy-counter case: staff has no time to type a
                # customer name during peak demand. Never search for an existing
                # tab by blank name (that would silently merge two unrelated
                # anonymous customers' bills together) — only look up an
                # existing tab when a name was actually given. (Quick-Sell
                # audit follow-up finding, 2026-07-19 — the previous
                # `and credit_recipient` gate meant a blank-name tab sale never
                # created a BarTab at all: payment_method_qs was already
                # correctly 'credit', but recipient='' on every line, so the
                # debt became an orphaned, unattributed credit transaction with
                # no tab, no PIN, no way to ever collect or look it up again.)
                bar_tab = (
                    BarTab.objects.filter(
                        business=user_profile.business, customer_name=credit_recipient,
                        status='OPEN', source='qs',
                    ).first()
                    if credit_recipient else None
                )
                if not bar_tab:
                    bar_tab = BarTab.create_with_credentials(
                        business=user_profile.business,
                        customer_name=credit_recipient,
                        served_by=request.user,
                        source='qs',
                    )
                    if not credit_recipient:
                        bar_tab.customer_name = f'Tab #{bar_tab.id}'
                        bar_tab.save(update_fields=['customer_name'])
                        credit_recipient = bar_tab.customer_name
                        # The transactions above were already saved with
                        # recipient='' (credit_recipient was blank at the time) —
                        # backfill now that the fallback name exists.
                        for _txn, _desc, _amt in tab_transactions:
                            _txn.recipient = credit_recipient
                            _txn.save(update_fields=['recipient'])
                for txn, desc, amt in tab_transactions:
                    BarTabEntry.objects.create(
                        tab=bar_tab,
                        transaction=txn,
                        description=desc,
                        amount=amt,
                    )
                # Ensure Customer record exists with phone so debt tracker can SMS them
                from .models import Customer as _Customer
                _cust = _Customer.objects.filter(
                    business=user_profile.business, name=credit_recipient
                ).first()
                if _cust is None:
                    _cust = _Customer.objects.create(
                        business=user_profile.business, name=credit_recipient,
                        credit_approved=True,
                    )
                if credit_phone and not _cust.phone:
                    _cust.phone = credit_phone
                    _cust.save(update_fields=["phone"])

                # Link customer FK on the tab so the drawer can retrieve their phone
                if bar_tab.customer_id != _cust.pk:
                    bar_tab.customer = _cust
                    bar_tab.save(update_fields=["customer"])

                # Issue or reuse master receipt for this tab so the customer has a live QR,
                # regardless of which counter (bar/kitchen/Quick Sell) rings up their next
                # item. Single source of truth — see core/tab_receipts.py.
                _tab_rcpt_token = None
                _tab_rcpt_number = None
                _tab_rcpt_id = None
                _tab_receipt_url = None
                _tab_master_rcpt = None
                _tab_freshly_linked = False
                try:
                    from .models import Receipt as _Rcpt
                    from core.tab_receipts import resolve_master_receipt
                    _tab_master_rcpt, _tab_freshly_linked = resolve_master_receipt(
                        user_profile.business, bar_tab
                    )
                    if _tab_master_rcpt:
                        _tab_rcpt_token  = _tab_master_rcpt.token
                        _tab_rcpt_number = _tab_master_rcpt.receipt_number
                        _tab_rcpt_id     = _tab_master_rcpt.id
                    else:
                        _new_rcpt = _Rcpt.issue(
                            business=user_profile.business,
                            lines=recorded,
                            payment_method='tab',
                            user=request.user,
                            customer_name=credit_recipient,
                            customer_phone=credit_phone,
                            meta={'tab_id': bar_tab.id},
                        )
                        _tab_rcpt_token  = _new_rcpt.token
                        _tab_rcpt_number = _new_rcpt.receipt_number
                        _tab_rcpt_id     = _new_rcpt.id
                    if _tab_rcpt_token:
                        _tab_receipt_url = request.build_absolute_uri(f'/r/{_tab_rcpt_token}/')
                except Exception:
                    pass

                # SMS to customer: brand-new tab, or freshly linked into an existing
                # tab/receipt from another counter (bar or kitchen) — no SMS on a plain
                # subsequent round against a receipt we already knew about.
                if (_tab_master_rcpt is None or _tab_freshly_linked) and _tab_receipt_url:
                    _sms_phone_raw = credit_phone or (_cust.phone if _cust else '')
                    if _sms_phone_raw:
                        try:
                            from .notifications import normalize_ke_phone, send_sms_notification
                            _sms_phone_qs = normalize_ke_phone(_sms_phone_raw)
                            if _sms_phone_qs:
                                if _tab_freshly_linked:
                                    _sms_qs = (
                                        f"Habari {credit_recipient},\n"
                                        f"{user_profile.business.name}: Bidhaa imeongezwa kwenye tab yako.\n"
                                        f"Angalia risiti iliyosasishwa: {_tab_receipt_url}"
                                    )
                                else:
                                    _sms_qs = (
                                        f"Habari {credit_recipient},\n"
                                        f"{user_profile.business.name}: Tab imefunguliwa — "
                                        f"KES {total:,.0f}.\n"
                                        f"Angalia risiti yako: {_tab_receipt_url}"
                                    )
                                send_sms_notification(_sms_qs, _sms_phone_qs)
                        except Exception:
                            pass

                success_data = json.dumps({
                    "items": recorded,
                    "total": total,
                    "payment_method": "tab",
                    "tab_customer": credit_recipient,
                    "receipt_token": _tab_rcpt_token,
                    "receipt_url": _tab_receipt_url,
                    "receipt_id": _tab_rcpt_id,
                })
                messages.success(
                    request,
                    _("Added to %(customer)s's tab: %(count)s item(s), KES %(total)s")
                    % {"customer": credit_recipient, "count": len(recorded), "total": f"{total:,.0f}"},
                )

            else:
                # ── DENI / regular cash / M-Pesa ─────────────────────────────
                # Auto-create Customer for credit (deni) sales so debt tracker finds them
                if payment_method_qs == "credit" and credit_recipient:
                    from .models import Customer as _Customer
                    cust_obj = _Customer.objects.filter(
                        business=user_profile.business, name=credit_recipient
                    ).first()
                    if cust_obj is None:
                        cust_obj = _Customer.objects.create(
                            business=user_profile.business, name=credit_recipient,
                            credit_approved=True,
                        )
                    if credit_phone and not cust_obj.phone:
                        cust_obj.phone = credit_phone
                        cust_obj.save(update_fields=["phone"])

                receipt_token = None
                receipt_number = None
                rcpt = None
                _qs_rcpt_reused = False  # True when lines appended to existing receipt
                try:
                    from .models import Receipt
                    from decimal import Decimal as _DecQS
                    rcpt_meta = {}
                    if payment_method_qs == "credit" and credit_recipient:
                        # Dedup: reuse today's receipt for this customer instead of
                        # issuing a new one (prevents duplicate SMS + multiple receipt links).
                        _existing_rcpt = Receipt.objects.filter(
                            business=user_profile.business,
                            customer_name__iexact=credit_recipient,
                            created_at__date=timezone.localdate(),
                        ).exclude(payment_method='statement').order_by('-created_at').first()
                        if _existing_rcpt:
                            _updated_lines = list(_existing_rcpt.lines) + recorded
                            _updated_total = sum(float(l.get('subtotal', 0)) for l in _updated_lines)
                            _existing_rcpt.lines = _updated_lines
                            _existing_rcpt.total = _DecQS(str(round(_updated_total, 2)))
                            _existing_rcpt.save(update_fields=['lines', 'total'])
                            rcpt = _existing_rcpt
                            receipt_token = _existing_rcpt.token
                            receipt_number = _existing_rcpt.receipt_number
                            _qs_rcpt_reused = True
                        else:
                            try:
                                from core.debt_views import _build_credit_receipt_meta
                                _cust_for_meta = _Customer.objects.filter(
                                    business=user_profile.business, name=credit_recipient
                                ).first()
                                if _cust_for_meta:
                                    rcpt_meta = _build_credit_receipt_meta(
                                        user_profile.business, _cust_for_meta, 'bar'
                                    )
                            except Exception:
                                pass
                            rcpt = Receipt.issue(
                                business=user_profile.business,
                                lines=recorded,
                                payment_method=payment_method_qs,
                                user=request.user,
                                customer_name=credit_recipient,
                                customer_phone=credit_phone,
                                meta=rcpt_meta,
                            )
                            receipt_token = rcpt.token
                            receipt_number = rcpt.receipt_number
                    else:
                        rcpt = Receipt.issue(
                            business=user_profile.business,
                            lines=recorded,
                            payment_method=payment_method_qs,
                            user=request.user,
                            customer_name="",
                            customer_phone="",
                            meta=rcpt_meta,
                        )
                        receipt_token = rcpt.token
                        receipt_number = rcpt.receipt_number
                except Exception:
                    pass

                # SMS confirmation: only for brand-new receipts (not when appended to existing)
                if payment_method_qs == "credit" and credit_phone and receipt_token and not _qs_rcpt_reused:
                    try:
                        from .notifications import normalize_ke_phone, send_sms_notification
                        from django.utils import timezone as _tz
                        normalized = normalize_ke_phone(credit_phone)
                        if normalized:
                            credit_window = user_profile.business.credit_window_days or 30
                            due_date = (_tz.now().date() + __import__('datetime').timedelta(days=credit_window)).strftime('%d %b %Y')
                            receipt_url_sms = request.build_absolute_uri(f'/r/{receipt_token}/')
                            sms_msg = (
                                f"Duka: {user_profile.business.name}\n"
                                f"Umenunua kwa deni: KES {total:,.0f}\n"
                                f"Tarehe ya malipo: {due_date}\n"
                                f"Risiti: {receipt_url_sms}"
                            )
                            send_sms_notification(sms_msg, normalized)
                    except Exception:
                        pass

                receipt_url = (
                    request.build_absolute_uri(f'/r/{receipt_token}/')
                    if receipt_token else None
                )
                success_data = json.dumps({
                    "items": recorded,
                    "total": total,
                    "payment_method": payment_method_raw,
                    "tab_customer": None,
                    "receipt_token": receipt_token,
                    "receipt_number": receipt_number,
                    "receipt_url": receipt_url,
                    "receipt_id": rcpt.id if rcpt else None,
                })
                messages.success(
                    request,
                    _("Sale recorded: %(item_count)s item(s), KES %(total)s")
                    % {"item_count": len(recorded), "total": f"{total:,.0f}"},
                )

    # Station scoping: kitchen-only staff belong on the Kitchen Board, not Quick Sell.
    _qs_show_bar, _qs_show_kitchen = _station_scope(user_profile)
    if _qs_show_kitchen and not _qs_show_bar:
        return redirect('kitchen_board')

    items_qs = list(
        Item.objects.filter(store__business=user_profile.business)
        .exclude(is_produce=True, produce_mode="BUNCH")  # greens render in their own board
        .exclude(is_keg=True)  # keg items render in the bar board
        .exclude(store__is_kitchen=True)  # kitchen items render on Kitchen Board only
        .select_related("store")
        .prefetch_related("portion_presets")
        .order_by("description")
    )

    if items_qs:
        item_ids = [item.id for item in items_qs]
        txn_aggregates = (
            Transaction.objects.filter(item_id__in=item_ids)
            .values("item_id")
            .annotate(total_qty=Sum("qty"))
        )
        balance_lookup = {
            agg["item_id"]: (agg["total_qty"] or 0) for agg in txn_aggregates
        }
    else:
        balance_lookup = {}

    # Annotate with pending restock request flag — mirrors stock_list()'s same
    # check, so Quick Sell staff can raise (and see) a restock request without
    # navigating away from the point-of-sale screen mid-shift, same as bar
    # board and kitchen board already can (Quick-Sell-module audit finding,
    # 2026-07-19 — QS was the only one of the three counters missing this).
    _qs_pending_restock_ids = set(
        StockRequest.objects.filter(
            business=user_profile.business,
            status__in=[StockRequest.STATUS_PENDING, StockRequest.STATUS_ORDERED],
            item_id__in=[i.id for i in items_qs],
        ).values_list('item_id', flat=True)
    )

    items = []
    for item in items_qs:
        txn_sum = balance_lookup.get(item.id, 0)
        balance = item.opening_bin_balance + txn_sum
        items.append(
            {
                "id": item.id,
                "description": item.description,
                "selling_price": item.selling_price,
                "balance": balance,
                "unit": item.unit,
                "store_id": item.store_id,
                "reorder_level": item.reorder_level,
                "is_produce": item.is_produce,
                "has_presets": len(item.portion_presets.all()) > 0,
                "volume_ml": item.volume_ml,
                "has_pending_restock": item.id in _qs_pending_restock_ids,
            }
        )

    # Exclude kitchen stores — no kitchen items appear in QS so the pill would show 0 items
    stores = Store.objects.filter(business=user_profile.business, is_kitchen=False)

    # Tab autocomplete: QS tabs only (separate lane from bar board tabs)
    open_tab_names = list(
        BarTab.objects.filter(business=user_profile.business, status='OPEN', source='qs')
        .values_list('customer_name', flat=True)
        .distinct()
    )

    return render(
        request,
        "core/quick_sell.html",
        {
            "items": items,
            "stores": stores,
            "success_data": success_data,
            "is_owner": (user_profile.is_owner_or_manager if user_profile else False),
            "open_tab_names": open_tab_names,
            "qs_items": items_qs,
        },
    )


@login_required
def next_material_no(request):
    """AJAX — returns the next available material_no for a given single-letter prefix.
    Used by the item form to auto-suggest G-01, G-02 … when category is Gin."""
    up = get_user_profile(request)
    if not up:
        return JsonResponse({'next': ''})
    prefix = (request.GET.get('prefix') or '').strip().upper()
    if not prefix or len(prefix) > 4 or not prefix[0].isalpha():
        return JsonResponse({'next': ''})
    # material_no is globally unique (no business scope), so query all items
    existing = Item.objects.filter(
        material_no__istartswith=prefix + '-',
    ).values_list('material_no', flat=True)

    max_num = 0
    for mn in existing:
        try:
            suffix = mn.split('-', 1)[1]
            if suffix.isdigit():
                max_num = max(max_num, int(suffix))
        except (IndexError, ValueError):
            pass

    return JsonResponse({'next': f'{prefix}-{max_num + 1:02d}'})


@login_required
def item_portion_presets(request, item_id):
    """AJAX — returns portion presets for any item with presets. Called by Quick Sell and Add Transaction."""
    user_profile = get_user_profile(request)
    item = get_object_or_404(Item, id=item_id, store__business=user_profile.business)

    presets = list(item.portion_presets.values(
        'id', 'label', 'price', 'quantity_consumed', 'display_order'
    ).order_by('display_order'))
    if not item.is_produce and not presets:
        return JsonResponse({'is_produce': False, 'presets': []})

    for p in presets:
        p['price'] = float(p['price'])
        p['quantity_consumed'] = float(p['quantity_consumed'])

    return JsonResponse({
        'is_produce': item.is_produce,
        'item_name': item.description,
        'unit': item.unit,
        'presets': presets,
        'balance': float(item.current_balance()),
    })


@login_required
@require_POST
def forecast_api(request):
    """
    POST /analytics/forecast/
    Body params (form or JSON):
        start_date  – YYYY-MM-DD
        end_date    – YYYY-MM-DD
        horizon     – int, days ahead (default 30)
        model       – 'ets' | 'regression' (default 'ets')
        item_id     – int or '' (optional)
    Returns JSON ready for Chart.js.
    """
    import json as _json

    profile = getattr(request.user, 'userprofile', None)
    if not profile or not profile.is_owner_or_manager:
        # Not a redirect (owner_or_manager_required's usual behavior) — this
        # is a JSON endpoint. The page that calls it (analytics.html's "Run
        # Forecast" button) is already owner/manager-gated, but the endpoint
        # itself had no gate of its own, so any staff member who knew/guessed
        # the URL could POST directly and pull a full revenue+profit history
        # and forecast, unfiltered by station.
        return JsonResponse({"error": "Owner or manager only"}, status=403)

    try:
        if request.content_type == "application/json":
            body = _json.loads(request.body)
        else:
            body = request.POST

        start_str = body.get("start_date", "")
        end_str = body.get("end_date", "")
        horizon = int(body.get("horizon", 30))
        model = body.get("model", "ets").lower()
        item_id = body.get("item_id") or None

        if not start_str or not end_str:
            return JsonResponse(
                {"error": "start_date and end_date are required."}, status=400
            )

        start = date.fromisoformat(start_str)
        end = date.fromisoformat(end_str)

        if start >= end:
            return JsonResponse(
                {"error": "start_date must be before end_date."}, status=400
            )

        horizon = max(1, min(horizon, 365))

        profile = request.user.userprofile
        business = profile.business

        from core.models import Transaction

        qs = Transaction.objects.filter(
            business=business,
            type="Issue",
            date__gte=start,
            date__lte=end,
        )
        if item_id:
            qs = qs.filter(item_id=item_id)

        if model == "regression":
            result = run_regression(qs, start, end, horizon)
        else:
            result = run_ets(qs, start, end, horizon)

        return JsonResponse(result)

    except Exception as exc:
        return JsonResponse({"error": str(exc)}, status=500)
