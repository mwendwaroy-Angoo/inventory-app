import os
import django
from datetime import datetime

# Setup Django environment
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'stockapp.settings')
django.setup()

from core.models import Store, Item, Transaction
import openpyxl

# Load the workbook
wb = openpyxl.load_workbook('Official Master Stock Control Sheet SF-Computer & Maintenance Stores REVISING.xlsx')
print("Excel file loaded successfully!")

# Step 1: Create Stores
stores = ['SF STORE', 'COMPUTER STORE', 'MAINTENANCE STORE']
for store_name in stores:
    Store.objects.get_or_create(name=store_name)
print("Stores created.")

# Step 2: Import Items from STOCK CONTROL SHEET
sheet = wb['STOCK CONTROL SHEET']

# Find the header row (row 3 based on your file)
headers = None
data_start_row = 4  # Data starts from row 4

for row in sheet.iter_rows(min_row=1, max_row=10, values_only=True):
    if row[1] == 'Material No':
        headers = row
        break

if not headers:
    print("Headers not found!")
    exit()

print("Headers found:", headers)

# Map column indices
col_material = headers.index('Material No')
col_desc = headers.index('Item Description')
col_unit = headers.index('Unit of Issue')
col_open_bin = headers.index('Opening S5 Bin Card Bal.')
col_open_phys = headers.index('Opening Physical Stock Count.')
col_reorder_qty = headers.index('Reorder Quantity')
col_reorder_level = headers.index('Reorder Levels')

items_created = 0
for row in sheet.iter_rows(min_row=data_start_row, values_only=True):
    if row[col_material] is None:
        continue  # Skip empty rows

    material_no = str(row[col_material]).strip()
    description = str(row[col_desc]).strip() if row[col_desc] else ""
    unit = str(row[col_unit]).strip() if row[col_unit] else ""

    open_bin = int(row[col_open_bin]) if row[col_open_bin] not in [None, ""] else 0
    open_phys = int(row[col_open_phys]) if row[col_open_phys] not in [None, ""] else 0
    reorder_qty = int(row[col_reorder_qty]) if row[col_reorder_qty] not in [None, ""] else 0
    reorder_level = int(row[col_reorder_level]) if row[col_reorder_level] not in [None, ""] else 0

    # All items seem to belong to SF STORE based on your file
    store = Store.objects.get(name='SF STORE')

    item, created = Item.objects.get_or_create(
        material_no=material_no,
        defaults={
            'store': store,
            'description': description,
            'unit': unit,
            'opening_bin_balance': open_bin,
            'opening_physical': open_phys,
            'reorder_quantity': reorder_qty,
            'reorder_level': reorder_level,
        }
    )
    if created:
        items_created += 1
    else:
        # Update if exists
        item.description = description
        item.unit = unit
        item.opening_bin_balance = open_bin
        item.opening_physical = open_phys
        item.reorder_quantity = reorder_qty
        item.reorder_level = reorder_level
        item.store = store
        item.save()

print(f"{items_created} new items created/updated.")

# Step 3: Import Transactions from SF STORE sheet
trans_sheet = wb['SF STORE']
trans_created = 0

for row in trans_sheet.iter_rows(min_row=3, values_only=True):  # Data starts row 3
    if not row[0] or not row[3]:  # Skip if no date or material no
        continue

    date_val = row[0]
    if isinstance(date_val, datetime):
        trans_date = date_val.date()
    else:
        trans_date = datetime(2025, 1, 1).date()  # fallback

    doc_no = str(row[1]) if row[1] else ""
    trans_type = str(row[2]).capitalize() if row[2] else "Issue"
    material_no = str(row[3]).strip()
    qty_raw = row[6]
    qty = int(qty_raw) if qty_raw else 0
    department = str(row[7]) if row[7] else ""

    # Determine sign: Receipt positive, others negative
    if trans_type == 'Receipt':
        quantity = qty
    else:
        quantity = -abs(qty)

    try:
        item = Item.objects.get(material_no=material_no)
    except Item.DoesNotExist:
        print(f"Item {material_no} not found, skipping transaction.")
        continue

    Transaction.objects.create(
        item=item,
        date=trans_date,
        doc_no=doc_no,
        type=trans_type,
        quantity=quantity,
        department=department
    )
    trans_created += 1

print(f"{trans_created} transactions imported from SF STORE.")

print("Import complete! Go to http://127.0.0.1:8000/admin to see your data.")